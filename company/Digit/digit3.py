# import streamlit as st
# import pandas as pd
# from io import BytesIO
# import base64
# import json
# import os
# from dotenv import load_dotenv
# import logging
# import re

# # Check if required packages are available
# try:
#     from openai import OpenAI
# except ImportError:
#     st.error("OpenAI package not found. Please install it using 'pip install openai'")
#     st.stop()

# try:
#     import PyPDF2
# except ImportError:
#     st.warning("PyPDF2 not found. PDF text extraction will use OpenAI vision only.")
#     PyPDF2 = None

# try:
#     import pdfplumber
# except ImportError:
#     st.warning("pdfplumber not found. PDF text extraction will use PyPDF2 or OpenAI vision only.")
#     pdfplumber = None

# try:
#     from openpyxl.cell import get_column_letter
# except ImportError:
#     get_column_letter = None

# # Load environment variables
# load_dotenv()

# # Configure logging
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# # Load OpenAI API key
# OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# if not OPENAI_API_KEY:
#     st.error("OPENAI_API_KEY environment variable not set. Please set it in your .env file or environment variables.")
#     st.stop()

# # Initialize OpenAI client
# try:
#     client = OpenAI(api_key=OPENAI_API_KEY)
# except Exception as e:
#     st.error(f"Failed to initialize OpenAI client: {str(e)}")
#     st.stop()

# # Embedded Formula Data 
# FORMULA_DATA = [
#     {"LOB": "TW", "SEGMENT": "1+5", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Bajaj, Digit, ICICI", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Bajaj, Digit, ICICI", "PO": "-3%", "REMARKS": "Payin Above 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR COMP + SAOD", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "All Fuel"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Bajaj, Digit, SBI", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Bajaj, Digit, SBI", "PO": "-3%", "REMARKS": "Payin Above 20%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Rest of Companies", "PO": "90% of Payin", "REMARKS": "Zuno -  21"},
#     {"LOB": "CV", "SEGMENT": "Upto 2.5 GVW", "INSURER": "Reliance, SBI, Tata", "PO": "-2%", "REMARKS": "NIL"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "BUS", "SEGMENT": "SCHOOL BUS", "INSURER": "TATA, Reliance, Digit, ICICI", "PO": "Less 2% of Payin", "REMARKS": "NIL"},
#     {"LOB": "BUS", "SEGMENT": "SCHOOL BUS", "INSURER": "Rest of Companies", "PO": "88% of Payin", "REMARKS": "NIL"},
#     {"LOB": "BUS", "SEGMENT": "STAFF BUS", "INSURER": "All Companies", "PO": "88% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "MISD", "SEGMENT": "Misd, Tractor", "INSURER": "All Companies", "PO": "88% of Payin", "REMARKS": "NIL"}
# ]


# def extract_text_from_pdf_file(pdf_bytes: bytes) -> str:
#     """Extract text from PDF using multiple methods"""
#     extracted_text = ""
    
#     # Method 1: Try pdfplumber first (most accurate)
#     if pdfplumber:
#         try:
#             with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
#                 for page in pdf.pages:
#                     page_text = page.extract_text()
#                     if page_text:
#                         extracted_text += page_text + "\n"
#             if extracted_text.strip():
#                 logger.info("PDF text extracted using pdfplumber")
#                 return extracted_text.strip()
#         except Exception as e:
#             logger.warning(f"pdfplumber extraction failed: {str(e)}")
    
#     # Method 2: Try PyPDF2 as fallback
#     if PyPDF2:
#         try:
#             pdf_reader = PyPDF2.PdfReader(BytesIO(pdf_bytes))
#             for page in pdf_reader.pages:
#                 page_text = page.extract_text()
#                 if page_text:
#                     extracted_text += page_text + "\n"
#             if extracted_text.strip():
#                 logger.info("PDF text extracted using PyPDF2")
#                 return extracted_text.strip()
#         except Exception as e:
#             logger.warning(f"PyPDF2 extraction failed: {str(e)}")
    
#     # Method 3: Return empty string if all methods fail
#     logger.warning("All PDF text extraction methods failed")
#     return ""


# def extract_text_from_file(file_bytes: bytes, filename: str, content_type: str) -> str:
#     """Extract text from uploaded file"""
#     file_extension = filename.split('.')[-1].lower() if '.' in filename else ''
#     file_type = content_type if content_type else file_extension

#     # Image-based extraction with enhanced OCR
#     image_extensions = ['png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff']
#     if file_extension in image_extensions or file_type.startswith('image/'):
#         image_base64 = base64.b64encode(file_bytes).decode('utf-8')
        
#         prompt = """Extract all insurance policy text accurately from this image using OCR.
#         Pay special attention to identifying:
#         - Agency/PB Clusters (these are locations)
#         - Agency/PB segments (these are vehicle segments)
#         - CD2 values (these are payin percentages - IGNORE CD1 completely)
#         - Policy codes like 1+1, 1+3, 1+5, SATP, TP
#         - Company names
#         - Any percentage values
#         - Any numerical data
#         - Table structure, including if values are blank under certain columns
#         - If CD1 columns exist, ignore them completely, only extract CD2 values
#         - If 1+1 CD2 (COMP) is blank but SATP CD2 (TP) has a value, note this relationship
#         Extract all text exactly as it appears in the image."""
            
#         response = client.chat.completions.create(
#             model="gpt-4o",
#             messages=[{
#                 "role": "user",
#                 "content": [
#                     {"type": "text", "text": prompt},
#                     {"type": "image_url", "image_url": {"url": f"data:image/{file_extension};base64,{image_base64}"}}
#                 ]
#             }],
#             temperature=0.1
#         )
#         return response.choices[0].message.content.strip()

#     # Enhanced PDF extraction
#     if file_extension == 'pdf':
#         # First try to extract text directly from PDF
#         pdf_text = extract_text_from_pdf_file(file_bytes)
        
#         if pdf_text and len(pdf_text.strip()) > 50:
#             # If we got good text extraction, use it
#             logger.info("Using direct PDF text extraction")
#             return pdf_text
#         else:
#             # Fallback to OpenAI vision for PDF
#             logger.info("Using OpenAI vision for PDF")
#             pdf_base64 = base64.b64encode(file_bytes).decode("utf-8")
            
#             prompt = """Extract insurance policy details from this PDF.
#             Focus on identifying Agency/PB Clusters (locations), segments, CD2 values (ignore CD1), and policy codes.
#             If 1+1 CD2 is blank but SATP CD2 has a value, note this relationship."""
                
#             response = client.chat.completions.create(
#                 model="gpt-4o",
#                 messages=[{
#                     "role": "user",
#                     "content": [
#                         {"type": "text", "text": prompt},
#                         {"type": "image_url", "image_url": {"url": f"data:application/pdf;base64,{pdf_base64}"}}
#                     ]
#                 }],
#                 temperature=0.1
#             )
#             return response.choices[0].message.content.strip()

#     # Text files
#     if file_extension == 'txt':
#         return file_bytes.decode('utf-8', errors='ignore')

#     # CSV files
#     if file_extension == 'csv':
#         df = pd.read_csv(BytesIO(file_bytes), encoding='utf-8', errors='ignore')
#         return df.to_string()

#     # Excel files - Fixed to handle binary Excel files
#     if file_extension in ['xlsx', 'xls', 'xlsm']:
#         try:
#             # Try to read with pandas
#             all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine='openpyxl')
#             dfs = []
#             for sheet_name, df_sheet in all_sheets.items():
#                 df_sheet["Source_Sheet"] = sheet_name
#                 dfs.append(df_sheet)
#             df = pd.concat(dfs, ignore_index=True, join="outer")
#             return df.to_string(index=False)
#         except Exception as e:
#             # If pandas fails, try with xlrd for older Excel files
#             try:
#                 all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine='xlrd')
#                 dfs = []
#                 for sheet_name, df_sheet in all_sheets.items():
#                     df_sheet["Source_Sheet"] = sheet_name
#                     dfs.append(df_sheet)
#                 df = pd.concat(dfs, ignore_index=True, join="outer")
#                 return df.to_string(index=False)
#             except Exception as e2:
#                 logger.error(f"Failed to read Excel file with both engines: {str(e)}, {str(e2)}")
#                 raise ValueError(f"Could not read Excel file: {filename}. Error: {str(e)}")

#     raise ValueError(f"Unsupported file type for {filename}")


# def clean_json_response(response_text: str) -> str:
#     """Clean and extract JSON from OpenAI response"""
#     # Remove markdown code blocks
#     cleaned = re.sub(r'```json\s*', '', response_text)
#     cleaned = re.sub(r'```\s*$', '', cleaned)
    
#     # Remove any text before the first [ or {
#     json_start = -1
#     for i, char in enumerate(cleaned):
#         if char in '[{':
#             json_start = i
#             break
    
#     if json_start != -1:
#         cleaned = cleaned[json_start:]
    
#     # Remove any text after the last ] or }
#     json_end = -1
#     for i in range(len(cleaned) - 1, -1, -1):
#         if cleaned[i] in ']}':
#             json_end = i + 1
#             break
    
#     if json_end != -1:
#         cleaned = cleaned[:json_end]
    
#     return cleaned.strip()


# def ensure_list_format(data) -> list:
#     """Ensure data is in list format"""
#     if isinstance(data, list):
#         return data
#     elif isinstance(data, dict):
#         return [data]  # Convert single object to list
#     else:
#         raise ValueError(f"Expected list or dict, got {type(data)}")


# def classify_payin(payin_str):
#     """
#     Converts Payin string (e.g., '50%') to float and classifies its range.
#     """
#     try:
#         # Handle various formats
#         payin_clean = str(payin_str).replace('%', '').replace(' ', '')
#         payin_value = float(payin_clean)
        
#         if payin_value <= 20:
#             category = "Payin Below 20%"
#         elif 21 <= payin_value <= 30:
#             category = "Payin 21% to 30%"
#         elif 31 <= payin_value <= 50:
#             category = "Payin 31% to 50%"
#         else:
#             category = "Payin Above 50%"
#         return payin_value, category
#     except (ValueError, TypeError):
#         logger.warning(f"Could not parse payin value: {payin_str}")
#         return 0.0, "Payin Below 20%"


# def apply_formula_directly(policy_data, company_name):
#     """
#     Apply formula rules directly using Python logic instead of OpenAI
#     """
#     calculated_data = []
    
#     for record in policy_data:
#         # Get policy details
#         segment = record.get('Segment', '') or ''
#         segment = str(segment)
#         policy_type = (record.get('Policy', '') or '').upper()
#         payin_value = record.get('Payin_Value', 0)
#         payin_category = record.get('Payin_Category', '')
        
#         # Determine LOB from segment
#         lob = ""
#         segment_upper = segment.upper()
        
#         if any(tw_keyword in segment_upper for tw_keyword in ['TW', '2W', 'TWO WHEELER', 'BIKE']):
#             lob = "TW"
#         elif any(car_keyword in segment_upper for car_keyword in ['PVT CAR', 'PRIVATE CAR', 'CAR']):
#             lob = "PVT CAR"
#         elif any(cv_keyword in segment_upper for cv_keyword in ['CV', 'COMMERCIAL', 'LCV', 'GVW', 'TN', 'PCV', 'GCV']):
#             lob = "CV"
#         elif 'BUS' in segment_upper:
#             lob = "BUS"
#         elif any(taxi_keyword in segment_upper for taxi_keyword in ['TAXI', 'PVT TAXI']):
#             lob = "TAXI"
#         elif any(misd_keyword in segment_upper for misd_keyword in ['MISD', 'TRACTOR', 'MISC', 'CRANE', 'GARBAGE']):
#             lob = "MISD"
#         else:
#             # Default to TW if uncertain
#             lob = "TW"
        
#         # Find matching formula rule
#         matched_rule = None
#         rule_explanation = ""
        
#         # Normalize company name for matching
#         company_normalized = company_name.upper().replace('GENERAL', '').replace('INSURANCE', '').strip()
        
#         for rule in FORMULA_DATA:
#             if rule["LOB"] != lob:
#                 continue
                
#             # Check if this rule applies to this segment and policy type
#             rule_segment = rule["SEGMENT"].upper()
            
#             # Segment matching logic
#             segment_match = False
#             if lob == "TW":
#                 if "1+5" in rule_segment and "1+5" in segment_upper:
#                     segment_match = True
#                 elif "SAOD + COMP" in rule_segment and (policy_type in ["COMP", "SAOD"]):
#                     segment_match = True
#                 elif "TP" in rule_segment and policy_type == "TP":
#                     segment_match = True
#             elif lob == "PVT CAR":
#                 if "COMP + SAOD" in rule_segment and (policy_type in ["COMP", "SAOD"]):
#                     segment_match = True
#                 elif "TP" in rule_segment and policy_type == "TP":
#                     segment_match = True
#             elif lob == "CV":
#                 if "GVW" in rule_segment or "PCV 3W" in rule_segment or "GCV 3W" in rule_segment:
#                     segment_match = True
#             elif lob == "BUS":
#                 if "SCHOOL BUS" in rule_segment and "SCHOOL" in segment_upper:
#                     segment_match = True
#                 elif "STAFF BUS" in rule_segment and "STAFF" in segment_upper:
#                     segment_match = True
#             elif lob == "TAXI":
#                 segment_match = True
#             elif lob == "MISD":
#                 segment_match = True
            
#             if not segment_match:
#                 continue
            
#             # Check company matching
#             insurers = [ins.strip().upper() for ins in rule["INSURER"].split(',')]
#             company_match = False
            
#             if "ALL COMPANIES" in insurers:
#                 company_match = True
#             elif "REST OF COMPANIES" in insurers:
#                 # Check if company is NOT in any specific company list for this LOB/SEGMENT
#                 is_in_specific_list = False
#                 for other_rule in FORMULA_DATA:
#                     if (other_rule["LOB"] == rule["LOB"] and 
#                         other_rule["SEGMENT"] == rule["SEGMENT"] and
#                         "REST OF COMPANIES" not in other_rule["INSURER"] and
#                         "ALL COMPANIES" not in other_rule["INSURER"]):
#                         other_insurers = [ins.strip().upper() for ins in other_rule["INSURER"].split(',')]
#                         if any(company_key in company_normalized for company_key in other_insurers):
#                             is_in_specific_list = True
#                             break
#                 if not is_in_specific_list:
#                     company_match = True
#             else:
#                 # Check specific company names
#                 for insurer in insurers:
#                     if insurer in company_normalized or company_normalized in insurer:
#                         company_match = True
#                         break
            
#             if not company_match:
#                 continue
            
#             # Check if remarks require payin category matching
#             remarks = rule.get("REMARKS", "")
            
#             if remarks == "NIL" or "NIL" in remarks.upper():
#                 # No payin category check needed - apply directly
#                 matched_rule = rule
#                 rule_explanation = f"Direct match: LOB={lob}, Segment={rule_segment}, Company={rule['INSURER']}, Policy={policy_type}"
#                 break
#             elif any(payin_keyword in remarks for payin_keyword in ["Payin Below", "Payin 21%", "Payin 31%", "Payin Above"]):
#                 # Need to match payin category
#                 if payin_category in remarks:
#                     matched_rule = rule
#                     rule_explanation = f"Payin category match: LOB={lob}, Segment={rule_segment}, Company={rule['INSURER']}, Payin={payin_category}"
#                     break
#             else:
#                 # Other remarks - apply directly
#                 matched_rule = rule
#                 rule_explanation = f"Other remarks match: LOB={lob}, Segment={rule_segment}, Company={rule['INSURER']}, Remarks={remarks}"
#                 break
        
#         # Calculate payout based on matched rule
#         if matched_rule:
#             po_formula = matched_rule["PO"]
#             calculated_payout = 0
            
#             if "90% of Payin" in po_formula:
#                 calculated_payout = payin_value * 0.9
#             elif "88% of Payin" in po_formula:
#                 calculated_payout = payin_value * 0.88
#             elif "Less 2% of Payin" in po_formula:
#                 calculated_payout = payin_value - 2
#             elif "-2%" in po_formula:
#                 calculated_payout = payin_value - 2
#             elif "-3%" in po_formula:
#                 calculated_payout = payin_value - 3
#             elif "-4%" in po_formula:
#                 calculated_payout = payin_value - 4
#             elif "-5%" in po_formula:
#                 calculated_payout = payin_value - 5
#             else:
#                 calculated_payout = payin_value  # Default to original payin
            
#             # Ensure non-negative result
#             calculated_payout = max(0, calculated_payout)
            
#             formula_used = po_formula
#         else:
#             # No rule matched - use default 90% of payin
#             calculated_payout = payin_value * 0.9
#             formula_used = "Default: 90% of Payin"
#             rule_explanation = f"No specific rule matched for LOB={lob}, Segment={segment}, Company={company_name}, Policy={policy_type}"
        
#         # Create result record in desired format
#         result_record = {
#     'Location': record.get('Location', 'N/A'),
#     'Segment': record.get('Segment', 'N/A'),
#     'Policy': record.get('Policy', 'N/A'),
#     'Payin': f"{int(payin_value)}%",
#     'Payout': f"{int(calculated_payout)}%",
#     'Remarks': record.get('Remarks', 'N/A') if record.get('Remarks') else 'NIL',
#     'Rule Explanation': rule_explanation
# }
        
#         calculated_data.append(result_record)
    
#     return calculated_data


# def process_files(policy_file_bytes, policy_filename, policy_content_type, company_name):
#     """Main processing function"""
#     try:
#         st.info("📄 Extracting text from policy file...")
        
#         # Extract text with enhanced intelligence
#         extracted_text = extract_text_from_file(policy_file_bytes, policy_filename, policy_content_type)

#         logger.info(f"Extracted text length: {len(extracted_text)}")

#         st.success(f"✅ Text extracted successfully! Policy: {len(extracted_text)} chars")

#         # Parse policy data with enhanced segment identification for Digit format
#         st.info("🧠 Parsing policy data with AI for Digit format...")
        
#         parse_prompt = f"""
#         Analyze the following insurance policy text from Digit and extract data according to these specific rules:

#         Company Name: {company_name}

#         IMPORTANT DIGIT FORMAT RULES:
#         - "Agency/PB Clusters" = Location (extract this as Location)
#         - "Agency/PB segment" = Segment (extract vehicle type)
#         - **CD2 = Payin percentage (COMPLETELY IGNORE CD1 - if CD1 columns exist, ignore them entirely)**
#         - **If 1+1 CD2 (COMP payin column) is blank, but SATP CD2 (TP column) has a value, use SATP CD2 value and set Policy="TP"**
        
#         SEGMENT TYPES FOR DIGIT:
#         - **Two Wheeler (TW)**: 1+5 (New TW), 1+1 (Old PVT car and TW), 1+3 (New PVT car)
#         - **PVT CAR**: with policy COMP + SAOD (where + means OR operator), also TP
#         - **CV**: GVW, PCV 3 Wheeler, GCV 3W
#         - **BUS**: School Bus, Staff Bus
#         - **TAXI**: includes PVT Taxi
#         - **MISD**: Tractor, Cranes, Garbage Vans
        
#         POLICY TYPES:
#         1. **COMP** (also known as Package/First Party): for Private Car, Two Wheelers, PCV (auto, bus), GCV, MISD
#         2. **SAOD**: comes with Private Car and Two Wheeler
#         3. **Third Party (TP)**: comes with Private Car, Two Wheeler, P.C.V, G.C.V & MISD
#         4. **1+1 means COMP**, **SATP means TP** for formula evaluation

#         CRITICAL CD1/CD2 HANDLING:
#         - **If CD1 columns are present, completely ignore them**
#         - **Only extract values from CD2 columns**
#         - **If 1+1 CD2 (COMP) is blank but SATP CD2 (TP) has value, use that value and set Policy="TP"**
#         - **If both have values, create separate records for each**

#         Extract into JSON records with these exact fields:
#         - "Location": from Agency/PB Clusters field
#         - "Segment": from Agency/PB segment field (include vehicle type details)
#         - "Policy": determine from codes and available values (1+1=COMP if has value, SATP=TP if has value)
#         - "Payin": from CD2 values only (convert to percentage format)
#         "Remarks": ALWAYS include additional information including vehicle makes (e.g. Hero/Honda), age info (e.g. Upto 180cc), validity, etc. If no extra info, set to "NIL"

#         EXAMPLE OUTPUT FORMAT:
#         [
#           {{
#             "Location": "RJ_Good",
#             "Segment": "TW 1+5", 
#             "Policy": "COMP",
#             "Payin": "30%",
#             "Remarks": "New Two Wheeler"
#           }},
#           {{
#             "Location": "RJ_Good",
#             "Segment": "TW SATP", 
#             "Policy": "TP",
#             "Payin": "25%",
#             "Remarks": "Third Party Only"
#           }}
#         ]

#         CRITICAL INSTRUCTIONS:
#         1. **IGNORE ALL CD1 VALUES COMPLETELY**
#         2. **ONLY USE CD2 VALUES FOR PAYIN**
#         3. **If 1+1 CD2 blank but SATP CD2 has value → use SATP value, set Policy="TP"**
#         4. **If both 1+1 CD2 and SATP CD2 have values → create 2 records (one COMP, one TP)**
#         5. **1+1 = COMP, SATP = TP for policy determination**
#         6. **PVT Taxi goes under TAXI segment**
#         7. **MISD includes Tractor, Cranes, Garbage Vans**

#         Text to analyze:
#         {extracted_text}
        
#         Return ONLY a valid JSON array, no other text.
#         """
        
#         response = client.chat.completions.create(
#             model="gpt-4o-mini",
#             messages=[
#                 {"role": "system", "content": "Extract insurance policy data as JSON array for Digit format. CRITICAL: Ignore CD1 completely, only use CD2. If 1+1 CD2 blank but SATP CD2 has value, use SATP and set Policy=TP. 1+1=COMP, SATP=TP. Return ONLY valid JSON array."},
#                 {"role": "user", "content": parse_prompt}
#             ],
#             temperature=0.1
#         )
        
#         parsed_json = response.choices[0].message.content.strip()
#         logger.info(f"Raw OpenAI response: {parsed_json[:500]}...")
        
#         # Clean the JSON response
#         cleaned_json = clean_json_response(parsed_json)
#         logger.info(f"Cleaned JSON: {cleaned_json[:500]}...")
        
#         try:
#             policy_data = json.loads(cleaned_json)
#             policy_data = ensure_list_format(policy_data)
#         except json.JSONDecodeError as e:
#             logger.error(f"JSON decode error: {str(e)}")
#             logger.error(f"Problematic JSON: {cleaned_json}")
            
#             # Fallback: create a basic structure
#             policy_data = [{
#                 "Location": "Unknown",
#                 "Segment": "TW 1+5",
#                 "Policy": "COMP", 
#                 "Payin": "30%",
#                 "Remarks": f"Error parsing data: {str(e)}"
#             }]

#         st.success(f"✅ Successfully parsed {len(policy_data)} policy records")

#         # Pre-classify Payin values
#         for record in policy_data:
#             try:
#                 payin_val, payin_cat = classify_payin(record.get('Payin', '0%'))
#                 record['Payin_Value'] = payin_val
#                 record['Payin_Category'] = payin_cat
#             except Exception as e:
#                 logger.warning(f"Error classifying payin for record: {record}, error: {str(e)}")
#                 record['Payin_Value'] = 30.0  # Default value
#                 record['Payin_Category'] = "Payin 21% to 30%"

#         # Apply formulas directly using Python logic
#         st.info("🧮 Applying formulas and calculating payouts...")
        
#         calculated_data = apply_formula_directly(policy_data, company_name)

#         st.success(f"✅ Successfully calculated {len(calculated_data)} records with payouts")

#         # Create Excel file with proper format
#         st.info("📊 Creating Excel file...")
        
#         df_calc = pd.DataFrame(calculated_data)
#         output = BytesIO()
        
#         # Use openpyxl engine to avoid binary format issues
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             # Write data starting from row 3 to leave space for headers
#             df_calc.to_excel(writer, sheet_name='Policy Data', startrow=2, index=False)
            
#             # Get the workbook and worksheet objects
#             workbook = writer.book
#             worksheet = writer.sheets['Policy Data']
            
#             # number of columns (ensure at least 1 to avoid merge errors)
#             cols_count = max(1, len(df_calc.columns))

#             # Set company name in row 1 (merged across columns)
#             worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_count)
#             company_cell = worksheet.cell(row=1, column=1, value=company_name)
#             try:
#                 company_cell.font = company_cell.font.copy(bold=True, size=14)
#                 company_cell.alignment = company_cell.alignment.copy(horizontal='center')
#             except Exception:
#                 pass
            
#             # Set subtitle in row 2 (merged across columns)
#             worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_count)
#             subtitle_cell = worksheet.cell(row=2, column=1, value='Policy Data with Payin and Calculated Payouts')
#             try:
#                 subtitle_cell.font = subtitle_cell.font.copy(bold=True, size=12)
#                 subtitle_cell.alignment = subtitle_cell.alignment.copy(horizontal='center')
#             except Exception:
#                 pass
            
#             # Apply header formatting to column headers (row 3)
#             for col_num, value in enumerate(df_calc.columns.values, 1):
#                 header_cell = worksheet.cell(row=3, column=col_num, value=value)
#                 try:
#                     header_cell.font = header_cell.font.copy(bold=True)
#                 except Exception:
#                     pass
            
#             # Auto-adjust column widths
#             for col_idx, column_name in enumerate(df_calc.columns, 1):
#                 max_length = 0
#                 # Use get_column_letter (safe) instead of accessing possibly-merged cell.column_letter
#                 if get_column_letter:
#                     column_letter = get_column_letter(col_idx)
#                 else:
#                     # fallback: use simple letter mapping for first 26 columns
#                     column_letter = chr(64 + col_idx) if col_idx <= 26 else str(col_idx)

#                 # Check header length
#                 max_length = max(max_length, len(str(column_name)))
                
#                 # Check data length
#                 for row_idx in range(4, len(df_calc) + 4):  # Start from row 4 (data starts there)
#                     cell_value = worksheet.cell(row=row_idx, column=col_idx).value
#                     if cell_value:
#                         max_length = max(max_length, len(str(cell_value)))
                
#                 adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
#                 try:
#                     worksheet.column_dimensions[column_letter].width = adjusted_width
#                 except Exception:
#                     # If column_dimensions fails for any reason, ignore and continue
#                     pass

#         output.seek(0)
#         excel_data = output.read()

#         return {
#             "extracted_text": extracted_text,
#             "parsed_data": policy_data,
#             "calculated_data": calculated_data,
#             "excel_data": excel_data,
#             "df_calc": df_calc
#         }

#     except Exception as e:
#         logger.error(f"Unexpected error: {str(e)}", exc_info=True)
#         raise Exception(f"An error occurred: {str(e)}")


# # Streamlit App
# def main():
#     st.set_page_config(
#         page_title="Insurance Policy Processing - Digit Format", 
#         page_icon="📋", 
#         layout="wide"
#     )
    
#     st.title("🏢 Insurance Policy Processing System - Digit Format")
#     st.markdown("---")
    
#     # Sidebar for inputs
#     with st.sidebar:
#         st.header("📁 File Upload")
        
#         # Company name input
#         company_name = st.text_input(
#             "Company Name", 
#             value="Digit",
#             help="Enter the insurance company name"
#         )
        
#         # Policy file upload
#         policy_file = st.file_uploader(
#             "📄 Upload Policy File",
#             type=['pdf', 'png', 'jpg', 'jpeg', 'txt', 'csv', 'xlsx', 'xls', 'xlsm'],
#             help="Upload your insurance policy document from Digit"
#         )
        
#         # Show Digit format mapping info
#         st.info("""
#         📊 **Digit Format Mapping:**
#         - Agency/PB Clusters → Location
#         - Agency/PB segment → Segment  
#         - **CD2 → Payin (CD1 IGNORED)**
#         - **1+1 CD2 blank + SATP CD2 value → Use SATP, Policy=TP**
#         - 1+5 → New TW COMP
#         - 1+1 → Old COMP (if has value)
#         - 1+3 → New PVT COMP
#         - SATP → TP
#         - PVT Taxi → TAXI
#         - MISD includes Tractor, Cranes, Garbage Vans
#         """)
        
#         # Process button
#         process_button = st.button(
#             "🚀 Process Digit Policy File", 
#             type="primary",
#             disabled=not policy_file
#         )
    
#     # Main content area
#     if not policy_file:
#         st.info("👆 Please upload a Digit policy file to begin processing.")
#         st.markdown("""
#         ### 📋 Digit Format Instructions:
#         1. **Company Name**: Enter the insurance company name (default: Digit)
#         2. **Policy File**: Upload the Digit document containing policy data
#         3. **Process**: Click the process button to extract data and calculate payouts
        
#         ### 🎯 Digit Format Key Features:
#         - **CD1/CD2 Handling**: **Completely ignores CD1 columns, only uses CD2 values**
#         - **Smart Blank Handling**: If 1+1 CD2 (COMP) is blank but SATP CD2 (TP) has value, uses SATP value and sets Policy=TP
#         - **Policy Code Recognition**: 
#           - 1+5 = New Two Wheeler COMP
#           - 1+1 = Old PVT car and TW COMP (if has value)
#           - 1+3 = New PVT car COMP
#           - SATP = TP (Third Party)
#         - **Segment Types**:
#           - TW: 1+5 (New), 1+1 (Old), includes bikes
#           - PVT CAR: COMP + SAOD (+ means OR), TP
#           - CV: GVW, PCV 3W, GCV 3W
#           - BUS: School Bus, Staff Bus
#           - TAXI: includes PVT Taxi
#           - MISD: Tractor, Cranes, Garbage Vans
        
#         ### 📊 Critical CD2 Logic:
#         - **If CD1 columns exist → IGNORE COMPLETELY**
#         - **Only extract from CD2 columns**
#         - **If 1+1 CD2 blank but SATP CD2 has value → Use SATP, Policy=TP**
#         - **If both have values → Create 2 records (COMP + TP)**
        
#         ### 📈 Output Format:
#         ```
#         Location    | Segment     | Policy | Payin | Payout | Remarks
#         RJ_Good     | TW 1+5      | COMP   | 30%   | 27%    | New Two Wheeler
#         RJ_Good     | TW SATP     | TP     | 25%   | 23%    | Third Party Only
#         ```
#         """)
#         return

#     if process_button:
#         try:
#             # Read file contents
#             policy_file_bytes = policy_file.read()
            
#             # Process files
#             with st.spinner("Processing Digit policy file... This may take a few moments."):
#                 results = process_files(
#                     policy_file_bytes, policy_file.name, policy_file.type,
#                     company_name
#                 )
            
#             st.success("🎉 Digit format processing completed successfully!")
            
#             # Display results in tabs
#             tab1, tab2, tab3, tab4 = st.tabs([
#                 "📊 Final Results", 
#                 "📁 Extracted Text", 
#                 "🧾 Parsed Data", 
#                 "📥 Download"
#             ])
            
#             with tab1:
#                 st.subheader("📊 Processed Digit Policy Data")
                
#                 # Display in the requested format
#                 st.markdown("### Location | Segment | Policy | Payin | Payout | Remarks")
#                 st.dataframe(results["df_calc"], use_container_width=True)
                
#                 # Summary statistics
#                 col1, col2, col3, col4 = st.columns(4)
#                 with col1:
#                     st.metric("Total Records", len(results["calculated_data"]))
#                 with col2:
#                     # Calculate average payin from the processed data
#                     payin_values = []
#                     for record in results["calculated_data"]:
#                         payin_str = record.get('Payin', '0%').replace('%', '')
#                         try:
#                             payin_values.append(float(payin_str))
#                         except:
#                             payin_values.append(0)
#                     avg_payin = sum(payin_values) / len(payin_values) if payin_values else 0
#                     st.metric("Avg Payin", f"{avg_payin:.1f}%")
#                 with col3:
#                     segments = set([r.get('Segment', 'N/A') for r in results["calculated_data"]])
#                     st.metric("Unique Segments", len(segments))
#                 with col4:
#                     st.metric("Company", company_name)
                
#                 # Show sample of formula application
#                 st.subheader("🔧 Sample Calculations")
#                 if results["calculated_data"]:
#                     sample_record = results["calculated_data"][0]
#                     st.write(f"**Example**: {sample_record.get('Segment')} policy with {sample_record.get('Payin')} payin → {sample_record.get('Payout')} payout (Explanation: {sample_record.get('Rule Explanation')})")
                
#                 # Show CD2 extraction info
#                 st.subheader("ℹ️ CD2 Processing Info")
#                 st.info("✅ CD1 columns were ignored. Only CD2 values used for Payin calculation.")
#                 st.info("✅ Blank 1+1 CD2 with SATP CD2 value → Used SATP and set Policy=TP")
            
#             with tab2:
#                 st.subheader("📁 Extracted Text from Digit Policy File")
#                 st.text_area(
#                     "Policy Text", 
#                     results["extracted_text"], 
#                     height=400,
#                     key="policy_text"
#                 )
                
#                 st.subheader("📊 Embedded Formula Rules")
#                 st.write("The system uses these formula rules:")
#                 df_formula = pd.DataFrame(FORMULA_DATA)
#                 st.dataframe(df_formula, use_container_width=True)
            
#             with tab3:
#                 st.subheader("🧾 Parsed Digit Policy Data (JSON)")
#                 st.json(results["parsed_data"])
                
#                 st.subheader("🧮 Calculated Data with Payouts")
#                 st.json(results["calculated_data"])
            
#             with tab4:
#                 st.subheader("📥 Download Results")
                
#                 # Excel download (fixed format)
#                 st.download_button(
#                     label="📊 Download Excel File (.xlsx)",
#                     data=results["excel_data"],
#                     file_name=f"{company_name}_digit_processed_policies.xlsx",
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                 )
                
#                 # JSON download
#                 json_data = json.dumps(results["calculated_data"], indent=2)
#                 st.download_button(
#                     label="📄 Download JSON Data",
#                     data=json_data,
#                     file_name=f"{company_name}_digit_processed_data.json",
#                     mime="application/json"
#                 )
                
#                 # CSV download
#                 csv_data = results["df_calc"].to_csv(index=False)
#                 st.download_button(
#                     label="📋 Download CSV File",
#                     data=csv_data,
#                     file_name=f"{company_name}_digit_processed_policies.csv",
#                     mime="text/csv"
#                 )
                
#                 st.success("✅ Files include Rule Explanation column")
#                 st.info("💡 The Excel file contains Digit processed data: Location | Segment | Policy | Payin | Payout | Remarks")
#                 st.info("🔍 CD1 ignored, CD2 used. Blank COMP with TP value handled correctly.")
                
#         except Exception as e:
#             st.error(f"❌ Error processing Digit files: {str(e)}")
#             st.exception(e)


# if __name__ == "__main__":
#     main()


# import streamlit as st
# import pandas as pd
# from io import BytesIO
# import base64
# import json
# import os
# from dotenv import load_dotenv
# import logging
# import re

# # Check if required packages are available
# try:
#     from openai import OpenAI
# except ImportError:
#     st.error("OpenAI package not found. Please install it using 'pip install openai'")
#     st.stop()

# try:
#     import PyPDF2
# except ImportError:
#     st.warning("PyPDF2 not found. PDF text extraction will use OpenAI vision only.")
#     PyPDF2 = None

# try:
#     import pdfplumber
# except ImportError:
#     st.warning("pdfplumber not found. PDF text extraction will use PyPDF2 or OpenAI vision only.")
#     pdfplumber = None

# try:
#     from openpyxl.cell import get_column_letter
# except ImportError:
#     get_column_letter = None

# # Load environment variables
# load_dotenv()

# # Configure logging
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# # Load OpenAI API key
# OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# if not OPENAI_API_KEY:
#     st.error("OPENAI_API_KEY environment variable not set. Please set it in your .env file or environment variables.")
#     st.stop()

# # Initialize OpenAI client
# try:
#     client = OpenAI(api_key=OPENAI_API_KEY)
# except Exception as e:
#     st.error(f"Failed to initialize OpenAI client: {str(e)}")
#     st.stop()

# # Embedded Formula Data
# FORMULA_DATA = [
#     {"LOB": "TW", "SEGMENT": "1+5", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Bajaj, Digit, ICICI", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Bajaj, Digit, ICICI", "PO": "-3%", "REMARKS": "Payin Above 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR COMP + SAOD", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "All Fuel"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Bajaj, Digit, SBI", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Bajaj, Digit, SBI", "PO": "-3%", "REMARKS": "Payin Above 20%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Rest of Companies", "PO": "90% of Payin", "REMARKS": "Zuno -  21"},
#     {"LOB": "CV", "SEGMENT": "Upto 2.5 GVW", "INSURER": "Reliance, SBI, Tata", "PO": "-2%", "REMARKS": "NIL"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "BUS", "SEGMENT": "SCHOOL BUS", "INSURER": "TATA, Reliance, Digit, ICICI", "PO": "Less 2% of Payin", "REMARKS": "NIL"},
#     {"LOB": "BUS", "SEGMENT": "SCHOOL BUS", "INSURER": "Rest of Companies", "PO": "88% of Payin", "REMARKS": "NIL"},
#     {"LOB": "BUS", "SEGMENT": "STAFF BUS", "INSURER": "All Companies", "PO": "88% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "MISD", "SEGMENT": "Misd, Tractor", "INSURER": "All Companies", "PO": "88% of Payin", "REMARKS": "NIL"}
# ]


# def extract_text_from_pdf_file(pdf_bytes: bytes) -> str:
#     """Extract text from PDF using multiple methods"""
#     extracted_text = ""
    
#     # Method 1: Try pdfplumber first (most accurate)
#     if pdfplumber:
#         try:
#             with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
#                 for page in pdf.pages:
#                     page_text = page.extract_text()
#                     if page_text:
#                         extracted_text += page_text + "\n"
#             if extracted_text.strip():
#                 logger.info("PDF text extracted using pdfplumber")
#                 return extracted_text.strip()
#         except Exception as e:
#             logger.warning(f"pdfplumber extraction failed: {str(e)}")
    
#     # Method 2: Try PyPDF2 as fallback
#     if PyPDF2:
#         try:
#             pdf_reader = PyPDF2.PdfReader(BytesIO(pdf_bytes))
#             for page in pdf_reader.pages:
#                 page_text = page.extract_text()
#                 if page_text:
#                     extracted_text += page_text + "\n"
#             if extracted_text.strip():
#                 logger.info("PDF text extracted using PyPDF2")
#                 return extracted_text.strip()
#         except Exception as e:
#             logger.warning(f"PyPDF2 extraction failed: {str(e)}")
    
#     # Method 3: Return empty string if all methods fail
#     logger.warning("All PDF text extraction methods failed")
#     return ""


# def extract_text_from_file(file_bytes: bytes, filename: str, content_type: str) -> str:
#     """Extract text from uploaded file"""
#     file_extension = filename.split('.')[-1].lower() if '.' in filename else ''
#     file_type = content_type if content_type else file_extension

#     # Image-based extraction with enhanced OCR
#     image_extensions = ['png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff']
#     if file_extension in image_extensions or file_type.startswith('image/'):
#         image_base64 = base64.b64encode(file_bytes).decode('utf-8')
        
#         prompt = """Extract all insurance policy text accurately from this image using OCR.
#         Pay special attention to identifying:
#         - Agency/PB Clusters (these are locations)
#         - Agency/PB segments (these are vehicle segments)
#         - CD2 values (these are payin percentages - IGNORE CD1 completely)
#         - Policy codes like 1+1, 1+3, 1+5, SATP, TP
#         - Company names
#         - Any percentage values
#         - Any numerical data
#         - Table structure, including if values are blank under certain columns
#         - If CD1 columns exist, ignore them completely, only extract CD2 values
#         - If 1+1 CD2 (COMP) is NIL but SATP CD2 (TP) has a value, note this relationship
#         Extract all text exactly as it appears in the image."""
            
#         response = client.chat.completions.create(
#             model="gpt-4o",
#             messages=[{
#                 "role": "user",
#                 "content": [
#                     {"type": "text", "text": prompt},
#                     {"type": "image_url", "image_url": {"url": f"data:image/{file_extension};base64,{image_base64}"}}
#                 ]
#             }],
#             temperature=0.1
#         )
#         return response.choices[0].message.content.strip()

#     # Enhanced PDF extraction
#     if file_extension == 'pdf':
#         # First try to extract text directly from PDF
#         pdf_text = extract_text_from_pdf_file(file_bytes)
        
#         if pdf_text and len(pdf_text.strip()) > 50:
#             # If we got good text extraction, use it
#             logger.info("Using direct PDF text extraction")
#             return pdf_text
#         else:
#             # Fallback to OpenAI vision for PDF
#             logger.info("Using OpenAI vision for PDF")
#             pdf_base64 = base64.b64encode(file_bytes).decode("utf-8")
            
#             prompt = """Extract insurance policy details from this PDF.
#             Focus on identifying Agency/PB Clusters (locations), segments, CD2 values (ignore CD1), and policy codes.
#             If 1+1 CD2 is NIL but SATP CD2 has a value, note this relationship."""
                
#             response = client.chat.completions.create(
#                 model="gpt-4o",
#                 messages=[{
#                     "role": "user",
#                     "content": [
#                         {"type": "text", "text": prompt},
#                         {"type": "image_url", "image_url": {"url": f"data:application/pdf;base64,{pdf_base64}"}}
#                     ]
#                 }],
#                 temperature=0.1
#             )
#             return response.choices[0].message.content.strip()

#     # Text files
#     if file_extension == 'txt':
#         return file_bytes.decode('utf-8', errors='ignore')

#     # CSV files
#     if file_extension == 'csv':
#         df = pd.read_csv(BytesIO(file_bytes), encoding='utf-8', errors='ignore')
#         return df.to_string()

#     # Excel files - Fixed to handle binary Excel files
#     if file_extension in ['xlsx', 'xls', 'xlsm']:
#         try:
#             # Try to read with pandas
#             all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine='openpyxl')
#             dfs = []
#             for sheet_name, df_sheet in all_sheets.items():
#                 df_sheet["Source_Sheet"] = sheet_name
#                 dfs.append(df_sheet)
#             df = pd.concat(dfs, ignore_index=True, join="outer")
#             return df.to_string(index=False)
#         except Exception as e:
#             # If pandas fails, try with xlrd for older Excel files
#             try:
#                 all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine='xlrd')
#                 dfs = []
#                 for sheet_name, df_sheet in all_sheets.items():
#                     df_sheet["Source_Sheet"] = sheet_name
#                     dfs.append(df_sheet)
#                 df = pd.concat(dfs, ignore_index=True, join="outer")
#                 return df.to_string(index=False)
#             except Exception as e2:
#                 logger.error(f"Failed to read Excel file with both engines: {str(e)}, {str(e2)}")
#                 raise ValueError(f"Could not read Excel file: {filename}. Error: {str(e)}")

#     raise ValueError(f"Unsupported file type for {filename}")


# def clean_json_response(response_text: str) -> str:
#     """Clean and extract JSON from OpenAI response"""
#     # Remove markdown code blocks
#     cleaned = re.sub(r'```json\s*', '', response_text)
#     cleaned = re.sub(r'```\s*$', '', cleaned)
    
#     # Remove any text before the first [ or {
#     json_start = -1
#     for i, char in enumerate(cleaned):
#         if char in '[{':
#             json_start = i
#             break
    
#     if json_start != -1:
#         cleaned = cleaned[json_start:]
    
#     # Remove any text after the last ] or }
#     json_end = -1
#     for i in range(len(cleaned) - 1, -1, -1):
#         if cleaned[i] in ']}':
#             json_end = i + 1
#             break
    
#     if json_end != -1:
#         cleaned = cleaned[:json_end]
    
#     return cleaned.strip()


# def ensure_list_format(data) -> list:
#     """Ensure data is in list format"""
#     if isinstance(data, list):
#         return data
#     elif isinstance(data, dict):
#         return [data]  # Convert single object to list
#     else:
#         raise ValueError(f"Expected list or dict, got {type(data)}")


# def classify_payin(payin_str):
#     """
#     Converts Payin string (e.g., '50%') to float and classifies its range.
#     """
#     try:
#         # Handle various formats
#         payin_clean = str(payin_str).replace('%', '').replace(' ', '')
#         payin_value = float(payin_clean)
        
#         if payin_value <= 20:
#             category = "Payin Below 20%"
#         elif 21 <= payin_value <= 30:
#             category = "Payin 21% to 30%"
#         elif 31 <= payin_value <= 50:
#             category = "Payin 31% to 50%"
#         else:
#             category = "Payin Above 50%"
#         return payin_value, category
#     except (ValueError, TypeError):
#         logger.warning(f"Could not parse payin value: {payin_str}")
#         return 0.0, "Payin Below 20%"


# def apply_formula_directly(policy_data, company_name):
#     """
#     Apply formula rules directly using Python logic instead of OpenAI
#     """
#     calculated_data = []
    
#     for record in policy_data:
#         # Get policy details
#         segment = record.get('Segment', '') or ''
#         segment = str(segment)
#         policy_type = (record.get('Policy', '') or '').upper()
#         payin_value = record.get('Payin_Value', 0)
#         payin_category = record.get('Payin_Category', '')
        
#         # Determine LOB from segment
#         lob = ""
#         segment_upper = segment.upper()
        
#         if any(tw_keyword in segment_upper for tw_keyword in ['TW', '2W', 'TWO WHEELER', 'BIKE']):
#             lob = "TW"
#         elif any(car_keyword in segment_upper for car_keyword in ['PVT CAR', 'PRIVATE CAR', 'CAR']):
#             lob = "PVT CAR"
#         elif any(cv_keyword in segment_upper for cv_keyword in ['CV', 'COMMERCIAL', 'LCV', 'GVW', 'TN', 'PCV', 'GCV']):
#             lob = "CV"
#         elif 'BUS' in segment_upper:
#             lob = "BUS"
#         elif any(taxi_keyword in segment_upper for taxi_keyword in ['TAXI', 'PVT TAXI']):
#             lob = "TAXI"
#         elif any(misd_keyword in segment_upper for misd_keyword in ['MISD', 'TRACTOR', 'MISC', 'CRANE', 'GARBAGE']):
#             lob = "MISD"
#         else:
#             # Default to TW if uncertain
#             lob = "TW"
        
#         # Find matching formula rule
#         matched_rule = None
#         rule_explanation = ""
        
#         # Normalize company name for matching
#         company_normalized = company_name.upper().replace('GENERAL', '').replace('INSURANCE', '').strip()
        
#         for rule in FORMULA_DATA:
#             if rule["LOB"] != lob:
#                 continue
                
#             # Check if this rule applies to this segment and policy type
#             rule_segment = rule["SEGMENT"].upper()
            
#             # Segment matching logic
#             segment_match = False
#             if lob == "TW":
#                 if "1+5" in rule_segment and "1+5" in segment_upper:
#                     segment_match = True
#                 elif "SAOD + COMP" in rule_segment and (policy_type in ["COMP", "SAOD"]):
#                     segment_match = True
#                 elif "TP" in rule_segment and policy_type == "TP":
#                     segment_match = True
#             elif lob == "PVT CAR":
#                 if "COMP + SAOD" in rule_segment and (policy_type in ["COMP", "SAOD"]):
#                     segment_match = True
#                 elif "TP" in rule_segment and policy_type == "TP":
#                     segment_match = True
#             elif lob == "CV":
#                 if "GVW" in rule_segment or "PCV 3W" in rule_segment or "GCV 3W" in rule_segment:
#                     segment_match = True
#             elif lob == "BUS":
#                 if "SCHOOL BUS" in rule_segment and "SCHOOL" in segment_upper:
#                     segment_match = True
#                 elif "STAFF BUS" in rule_segment and "STAFF" in segment_upper:
#                     segment_match = True
#             elif lob == "TAXI":
#                 segment_match = True
#             elif lob == "MISD":
#                 segment_match = True
            
#             if not segment_match:
#                 continue
            
#             # Check company matching
#             insurers = [ins.strip().upper() for ins in rule["INSURER"].split(',')]
#             company_match = False
            
#             if "ALL COMPANIES" in insurers:
#                 company_match = True
#             elif "REST OF COMPANIES" in insurers:
#                 # Check if company is NOT in any specific company list for this LOB/SEGMENT
#                 is_in_specific_list = False
#                 for other_rule in FORMULA_DATA:
#                     if (other_rule["LOB"] == rule["LOB"] and 
#                         other_rule["SEGMENT"] == rule["SEGMENT"] and
#                         "REST OF COMPANIES" not in other_rule["INSURER"] and
#                         "ALL COMPANIES" not in other_rule["INSURER"]):
#                         other_insurers = [ins.strip().upper() for ins in other_rule["INSURER"].split(',')]
#                         if any(company_key in company_normalized for company_key in other_insurers):
#                             is_in_specific_list = True
#                             break
#                 if not is_in_specific_list:
#                     company_match = True
#             else:
#                 # Check specific company names
#                 for insurer in insurers:
#                     if insurer in company_normalized or company_normalized in insurer:
#                         company_match = True
#                         break
            
#             if not company_match:
#                 continue
            
#             # Check if remarks require payin category matching
#             remarks = rule.get("REMARKS", "")
            
#             if remarks == "NIL" or "NIL" in remarks.upper():
#                 # No payin category check needed - apply directly
#                 matched_rule = rule
#                 rule_explanation = f"Direct match: LOB={lob}, Segment={rule_segment}, Company={rule['INSURER']}, Policy={policy_type}"
#                 break
#             elif any(payin_keyword in remarks for payin_keyword in ["Payin Below", "Payin 21%", "Payin 31%", "Payin Above"]):
#                 # Need to match payin category
#                 if payin_category in remarks:
#                     matched_rule = rule
#                     rule_explanation = f"Payin category match: LOB={lob}, Segment={rule_segment}, Company={rule['INSURER']}, Payin={payin_category}"
#                     break
#             else:
#                 # Other remarks - apply directly
#                 matched_rule = rule
#                 rule_explanation = f"Other remarks match: LOB={lob}, Segment={rule_segment}, Company={rule['INSURER']}, Remarks={remarks}"
#                 break
        
#         # Calculate payout based on matched rule
#         if matched_rule:
#             po_formula = matched_rule["PO"]
#             calculated_payout = 0
            
#             if "90% of Payin" in po_formula:
#                 calculated_payout = payin_value * 0.9
#             elif "88% of Payin" in po_formula:
#                 calculated_payout = payin_value * 0.88
#             elif "Less 2% of Payin" in po_formula:
#                 calculated_payout = payin_value - 2
#             elif "-2%" in po_formula:
#                 calculated_payout = payin_value - 2
#             elif "-3%" in po_formula:
#                 calculated_payout = payin_value - 3
#             elif "-4%" in po_formula:
#                 calculated_payout = payin_value - 4
#             elif "-5%" in po_formula:
#                 calculated_payout = payin_value - 5
#             else:
#                 calculated_payout = payin_value  # Default to original payin
            
#             # Ensure non-negative result
#             calculated_payout = max(0, calculated_payout)
            
#             formula_used = po_formula
#         else:
#             # No rule matched - use default 90% of payin
#             calculated_payout = payin_value * 0.9
#             formula_used = "Default: 90% of Payin"
#             rule_explanation = f"No specific rule matched for LOB={lob}, Segment={segment}, Company={company_name}, Policy={policy_type}"
        
#         # Create result record in desired format
#         result_record = {
#     'Location': record.get('Location', 'N/A'),
#     'Segment': record.get('Segment', 'N/A'),
#     'Policy': record.get('Policy', 'N/A'),
#     'Payin': f"{int(payin_value)}%",
#     'Payout': f"{int(calculated_payout)}%",
#     'Remarks': record.get('Remarks', 'N/A') if record.get('Remarks') else 'NIL',
#     'Rule Explanation': rule_explanation
# }
        
#         calculated_data.append(result_record)
    
#     return calculated_data


# def process_files(policy_file_bytes, policy_filename, policy_content_type, company_name):
#     """Main processing function"""
#     try:
#         st.info("📄 Extracting text from policy file...")
        
#         # Extract text with enhanced intelligence
#         extracted_text = extract_text_from_file(policy_file_bytes, policy_filename, policy_content_type)

#         logger.info(f"Extracted text length: {len(extracted_text)}")

#         st.success(f"✅ Text extracted successfully! Policy: {len(extracted_text)} chars")

#         # Parse policy data with enhanced segment identification for Digit format
#         st.info("🧠 Parsing policy data with AI for Digit format...")
        
#         parse_prompt = f"""
#         Analyze the following insurance policy text and extract data according to these specific rules:

#         Company Name: {company_name}

#         IMPORTANT FIELD MAPPING RULES:
#         - "Agency/PB Clusters" = Location (extract this as Location)
#         - "Agency/PB segment" = Segment (extract vehicle type like TW Bike, PVT CAR, etc.)
#         - "CD2" = Payin percentage (IGNORE CD1, only use CD2 values)
#         - If CD1 is present in the column, ignore the CD1 completely, just get the value of CD2
#         - If the value under 1+1 CD2 (or COMP payin column) is NIL, but beside it there is a column of SATP CD2 (or TP), and it contains the value, then consider that value for Payin and set Policy to "TP"
#         - Policy Type Codes:
#           * "1+5" = New Two Wheeler, COMP policy
#           * "1+1" = Old vehicle (PVT car or TW), COMP policy
#           * "1+3" = New PVT car, COMP policy
#           * "SATP" or "TP" = TP policy
#           * "COMP + SAOD" means COMP or SAOD (where + includes OR operator), set Policy to "COMP" or "SAOD" based on context
#           * If no specific code, assume COMP policy
#         - COMP also known as Package or first party, for Private Car, Two Wheelers, Passenger Carrying Vehicle (like auto, bus), Goods Carrying vehicles and MISD (Tractor, Cranes, Garbage Vans)
#         - SAOD comes with Private Car and Two Wheeler
#         - Third Party (TP) comes with Private Car, Two Wheeler, P.C.V, G.C.V & MISD
#         - 1+1 means COMP, SATP as TP for formula evaluation
#         - Any segment containing "SC/EV", "MC" will be classified under Two Wheelers (TW) segment.

#         SEGMENT IDENTIFICATION:
#         - Two Wheeler (TW): 1+5 for new TW, 1+1 for old PVT car and TW, 1+3 for new PVT car, and includes SC/EV, MC
#         - PVT car with policy COMP + SAOD (where + means OR)
#         - CV like GVW, PCV 3 Wheeler and GCV 3W
#         - Bus: School Bus and Staff Bus
#         - Taxi (PVT Taxi comes under Taxi)
#         - MISD (includes Tractor, Cranes, Garbage Vans)
#         - Look for vehicle types like: TW Bike, PVT CAR, CV, BUS, TAXI, MISD, Tractor, SC/EV, MC

#         Extract into JSON records with these exact fields:
#         - "Location": from Agency/PB Clusters field
#         - "Segment": from Agency/PB segment field (vehicle type, include details like 1+5, 1+1, etc. if present; map SC/EV and MC to TW)
#         - "Policy": determine from codes (1+5=COMP New TW, 1+1=COMP Old, 1+3=COMP New PVT, SATP=TP, COMP + SAOD=COMP or SAOD, default=COMP)
#         - "Payin": from CD2 values only (convert to percentage format, e.g. 0.30 → 30%, 25 → 25%)
#         - "Remarks": ALWAYS include additional information including vehicle makes (e.g. Hero/Honda), age info (e.g. Upto 180cc), validity, etc. If no extra info, set to "NIL"

#         EXAMPLE OUTPUT FORMAT:
#         [
#           {{
#             "Location": "RJ_Good",
#             "Segment": "TW Bike 1+5", 
#             "Policy": "COMP",
#             "Payin": "30%",
#             "Remarks": "Hero/Honda Upto 180cc"
#           }},
#           {{
#             "Location": "RJ_Good",
#             "Segment": "TW SC/EV", 
#             "Policy": "TP",
#             "Payin": "25%",
#             "Remarks": "Electric Vehicle"
#           }}
#         ]

#         Be very careful to:
#         1. ONLY use CD2 values for Payin (ignore CD1)
#         2. If COMP/1+1 CD2 is NIL, use SATP/TP CD2 if present and set Policy="TP"
#         3. Map policy codes correctly (1+5=COMP New TW, 1+1=COMP Old, 1+3=COMP New PVT, SATP=TP)
#         4. Extract Location from Agency/PB Clusters
#         5. Extract Segment from Agency/PB segment, including specifics like GVW, PCV, etc.; classify SC/EV and MC as TW
#         6. Convert all payin values to percentage format
#         7. Always return a valid JSON array, create separate records if a row has multiple policy types with values
#         8. PVT Taxi under Taxi
#         9. MISD includes Tractor, Cranes, Garbage Vans
#         10. ALWAYS fill Remarks with relevant info or "NIL" if none
#         Remember, if any remark contains NIL, then please consider the PO formula in the NIL remark row of the formula table
#         Text to analyze:
#         {extracted_text}
        
#         Return ONLY a valid JSON array, no other text.
#         """
#         response = client.chat.completions.create(
#             model="gpt-4o-mini",
#             messages=[
#                 {"role": "system", "content": "Extract insurance policy data as JSON array for Digit format. CRITICAL: Ignore CD1 completely, only use CD2. If 1+1 CD2 NIL but SATP CD2 has value, use SATP and set Policy=TP. 1+1=COMP, SATP=TP. Return ONLY valid JSON array."},
#                 {"role": "user", "content": parse_prompt}
#             ],
#             temperature=0.1
#         )
        
#         parsed_json = response.choices[0].message.content.strip()
#         logger.info(f"Raw OpenAI response: {parsed_json[:500]}...")
        
#         # Clean the JSON response
#         cleaned_json = clean_json_response(parsed_json)
#         logger.info(f"Cleaned JSON: {cleaned_json[:500]}...")
        
#         try:
#             policy_data = json.loads(cleaned_json)
#             policy_data = ensure_list_format(policy_data)
#         except json.JSONDecodeError as e:
#             logger.error(f"JSON decode error: {str(e)}")
#             logger.error(f"Problematic JSON: {cleaned_json}")
            
#             # Fallback: create a basic structure
#             policy_data = [{
#                 "Location": "Unknown",
#                 "Segment": "TW 1+5",
#                 "Policy": "COMP", 
#                 "Payin": "30%",
#                 "Remarks": f"Error parsing data: {str(e)}"
#             }]

#         st.success(f"✅ Successfully parsed {len(policy_data)} policy records")

#         # Pre-classify Payin values
#         for record in policy_data:
#             try:
#                 payin_val, payin_cat = classify_payin(record.get('Payin', '0%'))
#                 record['Payin_Value'] = payin_val
#                 record['Payin_Category'] = payin_cat
#             except Exception as e:
#                 logger.warning(f"Error classifying payin for record: {record}, error: {str(e)}")
#                 record['Payin_Value'] = 30.0  # Default value
#                 record['Payin_Category'] = "Payin 21% to 30%"

#         # Apply formulas directly using Python logic
#         st.info("🧮 Applying formulas and calculating payouts...")
        
#         calculated_data = apply_formula_directly(policy_data, company_name)

#         st.success(f"✅ Successfully calculated {len(calculated_data)} records with payouts")

#         # Create Excel file with proper format
#         st.info("📊 Creating Excel file...")
        
#         df_calc = pd.DataFrame(calculated_data)
#         output = BytesIO()
        
#         # Use openpyxl engine to avoid binary format issues
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             # Write data starting from row 3 to leave space for headers
#             df_calc.to_excel(writer, sheet_name='Policy Data', startrow=2, index=False)
            
#             # Get the workbook and worksheet objects
#             workbook = writer.book
#             worksheet = writer.sheets['Policy Data']
            
#             # number of columns (ensure at least 1 to avoid merge errors)
#             cols_count = max(1, len(df_calc.columns))

#             # Set company name in row 1 (merged across columns)
#             worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_count)
#             company_cell = worksheet.cell(row=1, column=1, value=company_name)
#             try:
#                 company_cell.font = company_cell.font.copy(bold=True, size=14)
#                 company_cell.alignment = company_cell.alignment.copy(horizontal='center')
#             except Exception:
#                 pass
            
#             # Set subtitle in row 2 (merged across columns)
#             worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_count)
#             subtitle_cell = worksheet.cell(row=2, column=1, value='Policy Data with Payin and Calculated Payouts')
#             try:
#                 subtitle_cell.font = subtitle_cell.font.copy(bold=True, size=12)
#                 subtitle_cell.alignment = subtitle_cell.alignment.copy(horizontal='center')
#             except Exception:
#                 pass
            
#             # Apply header formatting to column headers (row 3)
#             for col_num, value in enumerate(df_calc.columns.values, 1):
#                 header_cell = worksheet.cell(row=3, column=col_num, value=value)
#                 try:
#                     header_cell.font = header_cell.font.copy(bold=True)
#                 except Exception:
#                     pass
            
#             # Auto-adjust column widths
#             for col_idx, column_name in enumerate(df_calc.columns, 1):
#                 max_length = 0
#                 # Use get_column_letter (safe) instead of accessing possibly-merged cell.column_letter
#                 if get_column_letter:
#                     column_letter = get_column_letter(col_idx)
#                 else:
#                     # fallback: use simple letter mapping for first 26 columns
#                     column_letter = chr(64 + col_idx) if col_idx <= 26 else str(col_idx)

#                 # Check header length
#                 max_length = max(max_length, len(str(column_name)))
                
#                 # Check data length
#                 for row_idx in range(4, len(df_calc) + 4):  # Start from row 4 (data starts there)
#                     cell_value = worksheet.cell(row=row_idx, column=col_idx).value
#                     if cell_value:
#                         max_length = max(max_length, len(str(cell_value)))
                
#                 adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
#                 try:
#                     worksheet.column_dimensions[column_letter].width = adjusted_width
#                 except Exception:
#                     # If column_dimensions fails for any reason, ignore and continue
#                     pass

#         output.seek(0)
#         excel_data = output.read()

#         return {
#             "extracted_text": extracted_text,
#             "parsed_data": policy_data,
#             "calculated_data": calculated_data,
#             "excel_data": excel_data,
#             "df_calc": df_calc
#         }

#     except Exception as e:
#         logger.error(f"Unexpected error: {str(e)}", exc_info=True)
#         raise Exception(f"An error occurred: {str(e)}")


# # Streamlit App
# def main():
#     st.set_page_config(
#         page_title="Insurance Policy Processing - Digit Format", 
#         page_icon="📋", 
#         layout="wide"
#     )
    
#     st.title("🏢 Insurance Policy Processing System - Digit Format")
#     st.markdown("---")
    
#     # Sidebar for inputs
#     with st.sidebar:
#         st.header("📁 File Upload")
        
#         # Company name input
#         company_name = st.text_input(
#             "Company Name", 
#             value="Digit",
#             help="Enter the insurance company name"
#         )
        
#         # Policy file upload
#         policy_file = st.file_uploader(
#             "📄 Upload Policy File",
#             type=['pdf', 'png', 'jpg', 'jpeg', 'txt', 'csv', 'xlsx', 'xls', 'xlsm'],
#             help="Upload your insurance policy document from Digit"
#         )
        
#         # Show Digit format mapping info
#         st.info("""
#         📊 **Digit Format Mapping:**
#         - Agency/PB Clusters → Location
#         - Agency/PB segment → Segment  
#         - **CD2 → Payin (CD1 IGNORED)**
#         - **1+1 CD2 NIL + SATP CD2 value → Use SATP, Policy=TP**
#         - 1+5 → New TW COMP
#         - 1+1 → Old COMP (if has value)
#         - 1+3 → New PVT COMP
#         - SATP → TP
#         - PVT Taxi → TAXI
#         - MISD includes Tractor, Cranes, Garbage Vans
#         """)
        
#         # Process button
#         process_button = st.button(
#             "🚀 Process Digit Policy File", 
#             type="primary",
#             disabled=not policy_file
#         )
    
#     # Main content area
#     if not policy_file:
#         st.info("👆 Please upload a Digit policy file to begin processing.")
#         st.markdown("""
#         ### 📋 Digit Format Instructions:
#         1. **Company Name**: Enter the insurance company name (default: Digit)
#         2. **Policy File**: Upload the Digit document containing policy data
#         3. **Process**: Click the process button to extract data and calculate payouts
        
#         ### 🎯 Digit Format Key Features:
#         - **CD1/CD2 Handling**: **Completely ignores CD1 columns, only uses CD2 values**
#         - **Smart Blank Handling**: If 1+1 CD2 (COMP) is blank but SATP CD2 (TP) has value, uses SATP value and sets Policy=TP
#         - **Policy Code Recognition**: 
#           - 1+5 = New Two Wheeler COMP
#           - 1+1 = Old PVT car and TW COMP (if has value)
#           - 1+3 = New PVT car COMP
#           - SATP = TP (Third Party)
#         - **Segment Types**:
#           - TW: 1+5 (New), 1+1 (Old), includes bikes
#           - PVT CAR: COMP + SAOD (+ means OR), TP
#           - CV: GVW, PCV 3W, GCV 3W
#           - BUS: School Bus, Staff Bus
#           - TAXI: includes PVT Taxi
#           - MISD: Tractor, Cranes, Garbage Vans
        
#         ### 📊 Critical CD2 Logic:
#         - **If CD1 columns exist → IGNORE COMPLETELY**
#         - **Only extract from CD2 columns**
#         - **If 1+1 CD2 blank but SATP CD2 has value → Use SATP, Policy=TP**
#         - **If both have values → Create 2 records (COMP + TP)**
        
#         ### 📈 Output Format:
#         ```
#         Location    | Segment     | Policy | Payin | Payout | Remarks
#         RJ_Good     | TW 1+5      | COMP   | 30%   | 27%    | New Two Wheeler
#         RJ_Good     | TW SATP     | TP     | 25%   | 23%    | Third Party Only
#         ```
#         """)
#         return

#     if process_button:
#         try:
#             # Read file contents
#             policy_file_bytes = policy_file.read()
            
#             # Process files
#             with st.spinner("Processing Digit policy file... This may take a few moments."):
#                 results = process_files(
#                     policy_file_bytes, policy_file.name, policy_file.type,
#                     company_name
#                 )
            
#             st.success("🎉 Digit format processing completed successfully!")
            
#             # Display results in tabs
#             tab1, tab2, tab3, tab4 = st.tabs([
#                 "📊 Final Results", 
#                 "📁 Extracted Text", 
#                 "🧾 Parsed Data", 
#                 "📥 Download"
#             ])
            
#             with tab1:
#                 st.subheader("📊 Processed Digit Policy Data")
                
#                 # Display in the requested format
#                 st.markdown("### Location | Segment | Policy | Payin | Payout | Remarks")
#                 st.dataframe(results["df_calc"], use_container_width=True)
                
#                 # Summary statistics
#                 col1, col2, col3, col4 = st.columns(4)
#                 with col1:
#                     st.metric("Total Records", len(results["calculated_data"]))
#                 with col2:
#                     # Calculate average payin from the processed data
#                     payin_values = []
#                     for record in results["calculated_data"]:
#                         payin_str = record.get('Payin', '0%').replace('%', '')
#                         try:
#                             payin_values.append(float(payin_str))
#                         except:
#                             payin_values.append(0)
#                     avg_payin = sum(payin_values) / len(payin_values) if payin_values else 0
#                     st.metric("Avg Payin", f"{avg_payin:.1f}%")
#                 with col3:
#                     segments = set([r.get('Segment', 'N/A') for r in results["calculated_data"]])
#                     st.metric("Unique Segments", len(segments))
#                 with col4:
#                     st.metric("Company", company_name)
                
#                 # Show sample of formula application
#                 st.subheader("🔧 Sample Calculations")
#                 if results["calculated_data"]:
#                     sample_record = results["calculated_data"][0]
#                     st.write(f"**Example**: {sample_record.get('Segment')} policy with {sample_record.get('Payin')} payin → {sample_record.get('Payout')} payout (Explanation: {sample_record.get('Rule Explanation')})")
                
#                 # Show CD2 extraction info
#                 st.subheader("ℹ️ CD2 Processing Info")
#                 st.info("✅ CD1 columns were ignored. Only CD2 values used for Payin calculation.")
#                 st.info("✅ Blank 1+1 CD2 with SATP CD2 value → Used SATP and set Policy=TP")
            
#             with tab2:
#                 st.subheader("📁 Extracted Text from Digit Policy File")
#                 st.text_area(
#                     "Policy Text", 
#                     results["extracted_text"], 
#                     height=400,
#                     key="policy_text"
#                 )
                
#                 st.subheader("📊 Embedded Formula Rules")
#                 st.write("The system uses these formula rules:")
#                 df_formula = pd.DataFrame(FORMULA_DATA)
#                 st.dataframe(df_formula, use_container_width=True)
            
#             with tab3:
#                 st.subheader("🧾 Parsed Digit Policy Data (JSON)")
#                 st.json(results["parsed_data"])
                
#                 st.subheader("🧮 Calculated Data with Payouts")
#                 st.json(results["calculated_data"])
            
#             with tab4:
#                 st.subheader("📥 Download Results")
                
#                 # Excel download (fixed format)
#                 st.download_button(
#                     label="📊 Download Excel File (.xlsx)",
#                     data=results["excel_data"],
#                     file_name=f"{company_name}_digit_processed_policies.xlsx",
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                 )
                
#                 # JSON download
#                 json_data = json.dumps(results["calculated_data"], indent=2)
#                 st.download_button(
#                     label="📄 Download JSON Data",
#                     data=json_data,
#                     file_name=f"{company_name}_digit_processed_data.json",
#                     mime="application/json"
#                 )
                
#                 # CSV download
#                 csv_data = results["df_calc"].to_csv(index=False)
#                 st.download_button(
#                     label="📋 Download CSV File",
#                     data=csv_data,
#                     file_name=f"{company_name}_digit_processed_policies.csv",
#                     mime="text/csv"
#                 )
                
#                 st.success("✅ Files include Rule Explanation column")
#                 st.info("💡 The Excel file contains Digit processed data: Location | Segment | Policy | Payin | Payout | Remarks")
#                 st.info("🔍 CD1 ignored, CD2 used. Blank COMP with TP value handled correctly.")
                
#         except Exception as e:
#             st.error(f"❌ Error processing Digit files: {str(e)}")
#             st.exception(e)


# if __name__ == "__main__":
#     main()


# import streamlit as st
# import pandas as pd
# from io import BytesIO
# import base64
# import json
# import os
# from dotenv import load_dotenv
# import logging
# import re

# # Check if required packages are available
# try:
#     from openai import OpenAI
# except ImportError:
#     st.error("OpenAI package not found. Please install it using 'pip install openai'")
#     st.stop()

# try:
#     import PyPDF2
# except ImportError:
#     st.warning("PyPDF2 not found. PDF text extraction will use OpenAI vision only.")
#     PyPDF2 = None

# try:
#     import pdfplumber
# except ImportError:
#     st.warning("pdfplumber not found. PDF text extraction will use PyPDF2 or OpenAI vision only.")
#     pdfplumber = None

# try:
#     from openpyxl.cell import get_column_letter
# except ImportError:
#     get_column_letter = None

# # Load environment variables
# load_dotenv()

# # Configure logging
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# # Load OpenAI API key
# OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# if not OPENAI_API_KEY:
#     st.error("OPENAI_API_KEY environment variable not set. Please set it in your .env file or environment variables.")
#     st.stop()

# # Initialize OpenAI client
# try:
#     client = OpenAI(api_key=OPENAI_API_KEY)
# except Exception as e:
#     st.error(f"Failed to initialize OpenAI client: {str(e)}")
#     st.stop()

# # Embedded Formula Data
# FORMULA_DATA = [
#     {"LOB": "TW", "SEGMENT": "1+5", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Bajaj, Digit, ICICI", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Bajaj, Digit, ICICI", "PO": "-3%", "REMARKS": "Payin Above 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR COMP + SAOD", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "All Fuel"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Bajaj, Digit, SBI", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Bajaj, Digit, SBI", "PO": "-3%", "REMARKS": "Payin Above 20%"},
#     {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Rest of Companies", "PO": "90% of Payin", "REMARKS": "Zuno -  21"},
#     {"LOB": "CV", "SEGMENT": "Upto 2.5 GVW", "INSURER": "Reliance, SBI, Tata", "PO": "-2%", "REMARKS": "NIL"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "BUS", "SEGMENT": "SCHOOL BUS", "INSURER": "TATA, Reliance, Digit, ICICI", "PO": "Less 2% of Payin", "REMARKS": "NIL"},
#     {"LOB": "BUS", "SEGMENT": "SCHOOL BUS", "INSURER": "Rest of Companies", "PO": "88% of Payin", "REMARKS": "NIL"},
#     {"LOB": "BUS", "SEGMENT": "STAFF BUS", "INSURER": "All Companies", "PO": "88% of Payin", "REMARKS": "NIL"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
#     {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
#     {"LOB": "MISD", "SEGMENT": "Misd, Tractor", "INSURER": "All Companies", "PO": "88% of Payin", "REMARKS": "NIL"}
# ]


# def extract_text_from_pdf_file(pdf_bytes: bytes) -> str:
#     """Extract text from PDF using multiple methods"""
#     extracted_text = ""
    
#     # Method 1: Try pdfplumber first (most accurate)
#     if pdfplumber:
#         try:
#             with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
#                 for page in pdf.pages:
#                     page_text = page.extract_text()
#                     if page_text:
#                         extracted_text += page_text + "\n"
#             if extracted_text.strip():
#                 logger.info("PDF text extracted using pdfplumber")
#                 return extracted_text.strip()
#         except Exception as e:
#             logger.warning(f"pdfplumber extraction failed: {str(e)}")
    
#     # Method 2: Try PyPDF2 as fallback
#     if PyPDF2:
#         try:
#             pdf_reader = PyPDF2.PdfReader(BytesIO(pdf_bytes))
#             for page in pdf_reader.pages:
#                 page_text = page.extract_text()
#                 if page_text:
#                     extracted_text += page_text + "\n"
#             if extracted_text.strip():
#                 logger.info("PDF text extracted using PyPDF2")
#                 return extracted_text.strip()
#         except Exception as e:
#             logger.warning(f"PyPDF2 extraction failed: {str(e)}")
    
#     # Method 3: Return empty string if all methods fail
#     logger.warning("All PDF text extraction methods failed")
#     return ""


# def extract_text_from_file(file_bytes: bytes, filename: str, content_type: str) -> str:
#     """Extract text from uploaded file"""
#     file_extension = filename.split('.')[-1].lower() if '.' in filename else ''
#     file_type = content_type if content_type else file_extension

#     # Image-based extraction with enhanced OCR
#     image_extensions = ['png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff']
#     if file_extension in image_extensions or file_type.startswith('image/'):
#         image_base64 = base64.b64encode(file_bytes).decode('utf-8')
        
#         prompt = """Extract all insurance policy text accurately from this image using OCR.
#         Pay special attention to identifying:
#         - Agency/PB Clusters (these are locations)
#         - Agency/PB segments (these are vehicle segments)
#         - CD2 values (these are payin percentages - IGNORE CD1 completely)
#         - Policy codes like 1+1, 1+3, 1+5, SATP, TP
#         - Company names
#         - Any percentage values
#         - Any numerical data
#         - Table structure, including if values are blank under certain columns
#         - If CD1 columns exist, ignore them completely, only extract CD2 values
#         - If 1+1 CD2 (COMP) is NIL but SATP CD2 (TP) has a value, note this relationship
#         Extract all text exactly as it appears in the image."""
            
#         response = client.chat.completions.create(
#             model="gpt-4o",
#             messages=[{
#                 "role": "user",
#                 "content": [
#                     {"type": "text", "text": prompt},
#                     {"type": "image_url", "image_url": {"url": f"data:image/{file_extension};base64,{image_base64}"}}
#                 ]
#             }],
#             temperature=0.1
#         )
#         return response.choices[0].message.content.strip()

#     # Enhanced PDF extraction
#     if file_extension == 'pdf':
#         # First try to extract text directly from PDF
#         pdf_text = extract_text_from_pdf_file(file_bytes)
        
#         if pdf_text and len(pdf_text.strip()) > 50:
#             # If we got good text extraction, use it
#             logger.info("Using direct PDF text extraction")
#             return pdf_text
#         else:
#             # Fallback to OpenAI vision for PDF
#             logger.info("Using OpenAI vision for PDF")
#             pdf_base64 = base64.b64encode(file_bytes).decode("utf-8")
            
#             prompt = """Extract insurance policy details from this PDF.
#             Focus on identifying Agency/PB Clusters (locations), segments, CD2 values (ignore CD1), and policy codes.
#             If 1+1 CD2 is NIL but SATP CD2 has a value, note this relationship."""
                
#             response = client.chat.completions.create(
#                 model="gpt-4o",
#                 messages=[{
#                     "role": "user",
#                     "content": [
#                         {"type": "text", "text": prompt},
#                         {"type": "image_url", "image_url": {"url": f"data:application/pdf;base64,{pdf_base64}"}}
#                     ]
#                 }],
#                 temperature=0.1
#             )
#             return response.choices[0].message.content.strip()

#     # Text files
#     if file_extension == 'txt':
#         return file_bytes.decode('utf-8', errors='ignore')

#     # CSV files
#     if file_extension == 'csv':
#         df = pd.read_csv(BytesIO(file_bytes), encoding='utf-8', errors='ignore')
#         return df.to_string()

#     # Excel files - Fixed to handle binary Excel files
#     if file_extension in ['xlsx', 'xls', 'xlsm']:
#         try:
#             # Try to read with pandas
#             all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine='openpyxl')
#             dfs = []
#             for sheet_name, df_sheet in all_sheets.items():
#                 df_sheet["Source_Sheet"] = sheet_name
#                 dfs.append(df_sheet)
#             df = pd.concat(dfs, ignore_index=True, join="outer")
#             return df.to_string(index=False)
#         except Exception as e:
#             # If pandas fails, try with xlrd for older Excel files
#             try:
#                 all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine='xlrd')
#                 dfs = []
#                 for sheet_name, df_sheet in all_sheets.items():
#                     df_sheet["Source_Sheet"] = sheet_name
#                     dfs.append(df_sheet)
#                 df = pd.concat(dfs, ignore_index=True, join="outer")
#                 return df.to_string(index=False)
#             except Exception as e2:
#                 logger.error(f"Failed to read Excel file with both engines: {str(e)}, {str(e2)}")
#                 raise ValueError(f"Could not read Excel file: {filename}. Error: {str(e)}")

#     raise ValueError(f"Unsupported file type for {filename}")


# def clean_json_response(response_text: str) -> str:
#     """Clean and extract JSON from OpenAI response"""
#     # Remove markdown code blocks
#     cleaned = re.sub(r'```json\s*', '', response_text)
#     cleaned = re.sub(r'```\s*$', '', cleaned)
    
#     # Remove any text before the first [ or {
#     json_start = -1
#     for i, char in enumerate(cleaned):
#         if char in '[{':
#             json_start = i
#             break
    
#     if json_start != -1:
#         cleaned = cleaned[json_start:]
    
#     # Remove any text after the last ] or }
#     json_end = -1
#     for i in range(len(cleaned) - 1, -1, -1):
#         if cleaned[i] in ']}':
#             json_end = i + 1
#             break
    
#     if json_end != -1:
#         cleaned = cleaned[:json_end]
    
#     return cleaned.strip()


# def ensure_list_format(data) -> list:
#     """Ensure data is in list format"""
#     if isinstance(data, list):
#         return data
#     elif isinstance(data, dict):
#         return [data]  # Convert single object to list
#     else:
#         raise ValueError(f"Expected list or dict, got {type(data)}")


# def classify_payin(payin_str):
#     """
#     Converts Payin string (e.g., '50%') to float and classifies its range.
#     """
#     try:
#         # Handle various formats
#         payin_clean = str(payin_str).replace('%', '').replace(' ', '')
#         payin_value = float(payin_clean)
        
#         if payin_value <= 20:
#             category = "Payin Below 20%"
#         elif 21 <= payin_value <= 30:
#             category = "Payin 21% to 30%"
#         elif 31 <= payin_value <= 50:
#             category = "Payin 31% to 50%"
#         else:
#             category = "Payin Above 50%"
#         return payin_value, category
#     except (ValueError, TypeError):
#         logger.warning(f"Could not parse payin value: {payin_str}")
#         return 0.0, "Payin Below 20%"


# def apply_formula_directly(policy_data, company_name):
#     """
#     Apply formula rules directly using Python logic instead of OpenAI
#     """
#     calculated_data = []
    
#     for record in policy_data:
#         # Get policy details
#         segment = record.get('Segment', '') or ''
#         segment = str(segment)
#         policy_type = (record.get('Policy', '') or '').upper()
#         payin_value = record.get('Payin_Value', 0)
#         payin_category = record.get('Payin_Category', '')
        
#         # Determine LOB from segment
#         lob = ""
#         segment_upper = segment.upper()
        
#         if any(tw_keyword in segment_upper for tw_keyword in ['TW', '2W', 'TWO WHEELER', 'BIKE', 'SC/EV', 'MC']):
#             lob = "TW"
#         elif any(car_keyword in segment_upper for car_keyword in ['PVT CAR', 'PRIVATE CAR', 'CAR']):
#             lob = "PVT CAR"
#         elif any(cv_keyword in segment_upper for cv_keyword in ['CV', 'COMMERCIAL', 'LCV', 'GVW', 'TN', 'PCV', 'GCV']):
#             lob = "CV"
#         elif 'BUS' in segment_upper:
#             lob = "BUS"
#         elif any(taxi_keyword in segment_upper for taxi_keyword in ['TAXI', 'PVT TAXI']):
#             lob = "TAXI"
#         elif any(misd_keyword in segment_upper for misd_keyword in ['MISD', 'TRACTOR', 'MISC', 'CRANE', 'GARBAGE']):
#             lob = "MISD"
#         else:
#             # Default to TW if uncertain
#             lob = "TW"
        
#         # Normalize segment for TW with MC/SC/EV
#         normalized_segment = segment_upper
#         if any(mc_kw in segment_upper for mc_kw in ['MC', 'SC/EV']) and "TW" in segment_upper:
#             normalized_segment = "TW TP" if policy_type == "TP" else "TW"

#         # Find matching formula rule
#         matched_rule = None
#         rule_explanation = ""
        
#         # Normalize company name for matching
#         company_normalized = company_name.upper().replace('GENERAL', '').replace('INSURANCE', '').strip()
        
#         for rule in FORMULA_DATA:
#             if rule["LOB"] != lob:
#                 continue
                
#             # Check if this rule applies to this segment and policy type
#             rule_segment = rule["SEGMENT"].upper()
            
#             # Enhanced segment matching logic for TW
#             segment_match = False
#             if lob == "TW":
#                 if "1+5" in rule_segment and "1+5" in normalized_segment:
#                     segment_match = True
#                 elif "SAOD + COMP" in rule_segment and (policy_type in ["COMP", "SAOD"]):
#                     segment_match = True
#                 elif "TP" in rule_segment and policy_type == "TP" and any(tw_variant in normalized_segment for tw_variant in ["TW", "SC/EV", "MC"]):
#                     segment_match = True  # Match TW TP for TW variants including SC/EV and MC with TP policy
#                 elif "TP" in rule_segment and policy_type == "TP":
#                     segment_match = True
#             elif lob == "PVT CAR":
#                 if "COMP + SAOD" in rule_segment and (policy_type in ["COMP", "SAOD"]):
#                     segment_match = True
#                 elif "TP" in rule_segment and policy_type == "TP":
#                     segment_match = True
#             elif lob == "CV":
#                 if "GVW" in rule_segment or "PCV 3W" in rule_segment or "GCV 3W" in rule_segment:
#                     segment_match = True
#             elif lob == "BUS":
#                 if "SCHOOL BUS" in rule_segment and "SCHOOL" in normalized_segment:
#                     segment_match = True
#                 elif "STAFF BUS" in rule_segment and "STAFF" in normalized_segment:
#                     segment_match = True
#             elif lob == "TAXI":
#                 segment_match = True
#             elif lob == "MISD":
#                 segment_match = True
            
#             if not segment_match:
#                 continue
            
#             # Check company matching
#             insurers = [ins.strip().upper() for ins in rule["INSURER"].split(',')]
#             company_match = False
            
#             if "ALL COMPANIES" in insurers:
#                 company_match = True
#             elif "REST OF COMPANIES" in insurers:
#                 # Check if company is NOT in any specific company list for this LOB/SEGMENT
#                 is_in_specific_list = False
#                 for other_rule in FORMULA_DATA:
#                     if (other_rule["LOB"] == rule["LOB"] and 
#                         other_rule["SEGMENT"] == rule["SEGMENT"] and
#                         "REST OF COMPANIES" not in other_rule["INSURER"] and
#                         "ALL COMPANIES" not in other_rule["INSURER"]):
#                         other_insurers = [ins.strip().upper() for ins in other_rule["INSURER"].split(',')]
#                         if any(company_key in company_normalized for company_key in other_insurers):
#                             is_in_specific_list = True
#                             break
#                 if not is_in_specific_list:
#                     company_match = True
#             else:
#                 # Check specific company names
#                 for insurer in insurers:
#                     if insurer in company_normalized or company_normalized in insurer:
#                         company_match = True
#                         break
            
#             if not company_match:
#                 continue
            
#             # Check if remarks require payin category matching
#             remarks = rule.get("REMARKS", "")
            
#             if remarks == "NIL" or "NIL" in remarks.upper():
#                 # No payin category check needed - apply directly
#                 matched_rule = rule
#                 rule_explanation = f"Direct match: LOB={lob}, Segment={normalized_segment}, Company={rule['INSURER']}, Policy={policy_type}, No payin category check (NIL remarks)"
#                 break
#             elif any(payin_keyword in remarks for payin_keyword in ["Payin Below", "Payin 21%", "Payin 31%", "Payin Above"]):
#                 # Need to match payin category
#                 if payin_category in remarks:
#                     matched_rule = rule
#                     rule_explanation = f"Payin category match: LOB={lob}, Segment={normalized_segment}, Company={rule['INSURER']}, Payin={payin_category}, Rule remarks: {remarks}"
#                     break
#             else:
#                 # Other remarks - apply directly
#                 matched_rule = rule
#                 rule_explanation = f"Other remarks match: LOB={lob}, Segment={normalized_segment}, Company={rule['INSURER']}, Rule remarks: {remarks}"
#                 break
        
#         # Calculate payout based on matched rule
#         if matched_rule:
#             po_formula = matched_rule["PO"]
#             calculated_payout = 0
            
#             if "90% of Payin" in po_formula:
#                 calculated_payout = payin_value * 0.9
#             elif "88% of Payin" in po_formula:
#                 calculated_payout = payin_value * 0.88
#             elif "Less 2% of Payin" in po_formula:
#                 calculated_payout = payin_value - 2
#             elif "-2%" in po_formula:
#                 calculated_payout = payin_value - 2
#             elif "-3%" in po_formula:
#                 calculated_payout = payin_value - 3
#             elif "-4%" in po_formula:
#                 calculated_payout = payin_value - 4
#             elif "-5%" in po_formula:
#                 calculated_payout = payin_value - 5
#             else:
#                 calculated_payout = payin_value  # Default to original payin
            
#             # Ensure non-negative result
#             calculated_payout = max(0, calculated_payout)
            
#             formula_used = po_formula
#         else:
#             # No rule matched - use default 90% of payin
#             calculated_payout = payin_value * 0.9
#             formula_used = "Default: 90% of Payin"
#             rule_explanation = f"No specific rule matched for LOB={lob}, Segment={normalized_segment}, Company={company_name}, Policy={policy_type}, using default 90% of Payin"
        
#         # Create result record in desired format
#         result_record = {
#             'Location': record.get('Location', 'N/A'),
#             'Segment': record.get('Segment', 'N/A'),
#             'Policy': record.get('Policy', 'N/A'),
#             'Payin': f"{int(payin_value)}%",
#             'Payout': f"{int(calculated_payout)}%",
#             'Remarks': record.get('Remarks', 'N/A') if record.get('Remarks') else 'NIL',
#             'Rule Explanation': rule_explanation
#         }
        
#         calculated_data.append(result_record)
    
#     return calculated_data
# def process_files(policy_file_bytes, policy_filename, policy_content_type, company_name):
#     """Main processing function"""
#     try:
#         st.info("📄 Extracting text from policy file...")
        
#         # Extract text with enhanced intelligence
#         extracted_text = extract_text_from_file(policy_file_bytes, policy_filename, policy_content_type)

#         logger.info(f"Extracted text length: {len(extracted_text)}")

#         st.success(f"✅ Text extracted successfully! Policy: {len(extracted_text)} chars")

#         # Parse policy data with enhanced segment identification for Digit format
#         st.info("🧠 Parsing policy data with AI for Digit format...")
        
#         parse_prompt = f"""
#         Analyze the following insurance policy text and extract data according to these specific rules:

#         Company Name: {company_name}

#         IMPORTANT FIELD MAPPING RULES:
#         - "Agency/PB Clusters" = Location (extract this as Location)
#         - "Agency/PB segment" = Segment (extract vehicle type like TW Bike, PVT CAR, etc.)
#         - "CD2" = Payin percentage (IGNORE CD1, only use CD2 values)
#         - If CD1 is present in the column, ignore the CD1 completely, just get the value of CD2
#         - If the value under 1+1 CD2 (or COMP payin column) is blank, but beside it there is a column of SATP CD2 (or TP), and it contains the value, then consider that value for Payin and set Policy to "TP"
#         - Policy Type Codes:
#           * "1+5" = New Two Wheeler, COMP policy
#           * "1+1" = Old vehicle (PVT car or TW), COMP policy
#           * "1+3" = New PVT car, COMP policy
#           * "SATP" or "TP" = TP policy
#           * "COMP + SAOD" means COMP or SAOD (where + includes OR operator), set Policy to "COMP" or "SAOD" based on context
#           * If no specific code, assume COMP policy
#         - COMP also known as Package or first party, for Private Car, Two Wheelers, Passenger Carrying Vehicle (like auto, bus), Goods Carrying vehicles and MISD (Tractor, Cranes, Garbage Vans)
#         - SAOD comes with Private Car and Two Wheeler
#         - Third Party (TP) comes with Private Car, Two Wheelers, P.C.V, G.C.V & MISD
#         - 1+1 means COMP, SATP as TP for formula evaluation
#         - Any segment containing "SC/EV" or "MC" will be classified under Two Wheelers (TW) segment.

#         SEGMENT IDENTIFICATION:
#         - Two Wheeler (TW): 1+5 for new TW, 1+1 for old PVT car and TW, 1+3 for new PVT car, and includes SC/EV and MC
#         - PVT car with policy COMP + SAOD (where + means OR)
#         - CV like GVW, PCV 3 Wheeler and GCV 3W
#         - Bus: School Bus and Staff Bus
#         - Taxi (PVT Taxi comes under Taxi)
#         - MISD (includes Tractor, Cranes, Garbage Vans)
#         - Look for vehicle types like: TW Bike, PVT CAR, CV, BUS, TAXI, MISD, Tractor, SC/EV, MC

#         Extract into JSON records with these exact fields:
#         - "Location": from Agency/PB Clusters field
#         - "Segment": from Agency/PB segment field (vehicle type, include details like 1+5, 1+1, etc. if present; map SC/EV and MC to TW)
#         - "Policy": determine from codes (1+5=COMP New TW, 1+1=COMP Old, 1+3=COMP New PVT, SATP=TP, COMP + SAOD=COMP or SAOD, default=COMP)
#         - "Payin": from CD2 values only (convert to percentage format, e.g. 0.30 → 30%, 25 → 25%)
#         - "Remarks": ALWAYS include additional information including vehicle makes (e.g. Hero/Honda), age info (e.g. Upto 180cc), validity, etc. If no extra info, set to "NIL"

#         EXAMPLE OUTPUT FORMAT:
#         [
#           {{
#             "Location": "RJ_Good",
#             "Segment": "TW MC",
#             "Policy": "TP",
#             "Payin": "25%",
#             "Remarks": "Hero/Honda Upto 180cc"
#           }}
#         ]

#         Be very careful to:
#         1. ONLY use CD2 values for Payin (ignore CD1)
#         2. If COMP/1+1 CD2 is blank, use SATP/TP CD2 if present and set Policy="TP"
#         3. Map policy codes correctly (1+5=COMP New TW, 1+1=COMP Old, 1+3=COMP New PVT, SATP=TP)
#         4. Extract Location from Agency/PB Clusters
#         5. Extract Segment from Agency/PB segment, including specifics like GVW, PCV, etc.; map SC/EV and MC to TW
#         6. Convert all payin values to percentage format
#         7. Always return a valid JSON array, create separate records if a row has multiple policy types with values
#         8. PVT Taxi under Taxi
#         9. MISD includes Tractor, Cranes, Garbage Vans
#         10. ALWAYS fill Remarks with relevant info or "NIL" if none
#         Remember, if any remark contains NIL, then please consider the PO formula in the NIL remark row of the formula table
#         Text to analyze:
#         {extracted_text}
        
#         Return ONLY a valid JSON array, no other text.
#         """
#         response = client.chat.completions.create(
#             model="gpt-4o-mini",
#             messages=[
#                 {"role": "system", "content": "Extract insurance policy data as JSON array for Digit format. CRITICAL: Ignore CD1 completely, only use CD2. If 1+1 CD2 NIL but SATP CD2 has value, use SATP and set Policy=TP. 1+1=COMP, SATP=TP. Return ONLY valid JSON array."},
#                 {"role": "user", "content": parse_prompt}
#             ],
#             temperature=0.1
#         )
        
#         parsed_json = response.choices[0].message.content.strip()
#         logger.info(f"Raw OpenAI response: {parsed_json[:500]}...")
        
#         # Clean the JSON response
#         cleaned_json = clean_json_response(parsed_json)
#         logger.info(f"Cleaned JSON: {cleaned_json[:500]}...")
        
#         try:
#             policy_data = json.loads(cleaned_json)
#             policy_data = ensure_list_format(policy_data)
#         except json.JSONDecodeError as e:
#             logger.error(f"JSON decode error: {str(e)}")
#             logger.error(f"Problematic JSON: {cleaned_json}")
            
#             # Fallback: create a basic structure
#             policy_data = [{
#                 "Location": "Unknown",
#                 "Segment": "TW 1+5",
#                 "Policy": "COMP", 
#                 "Payin": "30%",
#                 "Remarks": f"Error parsing data: {str(e)}"
#             }]

#         st.success(f"✅ Successfully parsed {len(policy_data)} policy records")

#         # Pre-classify Payin values
#         for record in policy_data:
#             try:
#                 payin_val, payin_cat = classify_payin(record.get('Payin', '0%'))
#                 record['Payin_Value'] = payin_val
#                 record['Payin_Category'] = payin_cat
#             except Exception as e:
#                 logger.warning(f"Error classifying payin for record: {record}, error: {str(e)}")
#                 record['Payin_Value'] = 30.0  # Default value
#                 record['Payin_Category'] = "Payin 21% to 30%"

#         # Apply formulas directly using Python logic
#         st.info("🧮 Applying formulas and calculating payouts...")
        
#         calculated_data = apply_formula_directly(policy_data, company_name)

#         st.success(f"✅ Successfully calculated {len(calculated_data)} records with payouts")

#         # Create Excel file with proper format
#         st.info("📊 Creating Excel file...")
        
#         df_calc = pd.DataFrame(calculated_data)
#         output = BytesIO()
        
#         # Use openpyxl engine to avoid binary format issues
#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             # Write data starting from row 3 to leave space for headers
#             df_calc.to_excel(writer, sheet_name='Policy Data', startrow=2, index=False)
            
#             # Get the workbook and worksheet objects
#             workbook = writer.book
#             worksheet = writer.sheets['Policy Data']
            
#             # number of columns (ensure at least 1 to avoid merge errors)
#             cols_count = max(1, len(df_calc.columns))

#             # Set company name in row 1 (merged across columns)
#             worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_count)
#             company_cell = worksheet.cell(row=1, column=1, value=company_name)
#             try:
#                 company_cell.font = company_cell.font.copy(bold=True, size=14)
#                 company_cell.alignment = company_cell.alignment.copy(horizontal='center')
#             except Exception:
#                 pass
            
#             # Set subtitle in row 2 (merged across columns)
#             worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_count)
#             subtitle_cell = worksheet.cell(row=2, column=1, value='Policy Data with Payin and Calculated Payouts')
#             try:
#                 subtitle_cell.font = subtitle_cell.font.copy(bold=True, size=12)
#                 subtitle_cell.alignment = subtitle_cell.alignment.copy(horizontal='center')
#             except Exception:
#                 pass
            
#             # Apply header formatting to column headers (row 3)
#             for col_num, value in enumerate(df_calc.columns.values, 1):
#                 header_cell = worksheet.cell(row=3, column=col_num, value=value)
#                 try:
#                     header_cell.font = header_cell.font.copy(bold=True)
#                 except Exception:
#                     pass
            
#             # Auto-adjust column widths
#             for col_idx, column_name in enumerate(df_calc.columns, 1):
#                 max_length = 0
#                 # Use get_column_letter (safe) instead of accessing possibly-merged cell.column_letter
#                 if get_column_letter:
#                     column_letter = get_column_letter(col_idx)
#                 else:
#                     # fallback: use simple letter mapping for first 26 columns
#                     column_letter = chr(64 + col_idx) if col_idx <= 26 else str(col_idx)

#                 # Check header length
#                 max_length = max(max_length, len(str(column_name)))
                
#                 # Check data length
#                 for row_idx in range(4, len(df_calc) + 4):  # Start from row 4 (data starts there)
#                     cell_value = worksheet.cell(row=row_idx, column=col_idx).value
#                     if cell_value:
#                         max_length = max(max_length, len(str(cell_value)))
                
#                 adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
#                 try:
#                     worksheet.column_dimensions[column_letter].width = adjusted_width
#                 except Exception:
#                     # If column_dimensions fails for any reason, ignore and continue
#                     pass

#         output.seek(0)
#         excel_data = output.read()

#         return {
#             "extracted_text": extracted_text,
#             "parsed_data": policy_data,
#             "calculated_data": calculated_data,
#             "excel_data": excel_data,
#             "df_calc": df_calc
#         }

#     except Exception as e:
#         logger.error(f"Unexpected error: {str(e)}", exc_info=True)
#         raise Exception(f"An error occurred: {str(e)}")


# # Streamlit App
# def main():
#     st.set_page_config(
#         page_title="Insurance Policy Processing - Digit Format", 
#         page_icon="📋", 
#         layout="wide"
#     )
    
#     st.title("🏢 Insurance Policy Processing System - Digit Format")
#     st.markdown("---")
    
#     # Sidebar for inputs
#     with st.sidebar:
#         st.header("📁 File Upload")
        
#         # Company name input
#         company_name = st.text_input(
#             "Company Name", 
#             value="Digit",
#             help="Enter the insurance company name"
#         )
        
#         # Policy file upload
#         policy_file = st.file_uploader(
#             "📄 Upload Policy File",
#             type=['pdf', 'png', 'jpg', 'jpeg', 'txt', 'csv', 'xlsx', 'xls', 'xlsm'],
#             help="Upload your insurance policy document from Digit"
#         )
        
#         # Show Digit format mapping info
#         st.info("""
#         📊 **Digit Format Mapping:**
#         - Agency/PB Clusters → Location
#         - Agency/PB segment → Segment  
#         - **CD2 → Payin (CD1 IGNORED)**
#         - **1+1 CD2 NIL + SATP CD2 value → Use SATP, Policy=TP**
#         - 1+5 → New TW COMP
#         - 1+1 → Old COMP (if has value)
#         - 1+3 → New PVT COMP
#         - SATP → TP
#         - PVT Taxi → TAXI
#         - MISD includes Tractor, Cranes, Garbage Vans
#         """)
        
#         # Process button
#         process_button = st.button(
#             "🚀 Process Digit Policy File", 
#             type="primary",
#             disabled=not policy_file
#         )
    
#     # Main content area
#     if not policy_file:
#         st.info("👆 Please upload a Digit policy file to begin processing.")
#         st.markdown("""
#         ### 📋 Digit Format Instructions:
#         1. **Company Name**: Enter the insurance company name (default: Digit)
#         2. **Policy File**: Upload the Digit document containing policy data
#         3. **Process**: Click the process button to extract data and calculate payouts
        
#         ### 🎯 Digit Format Key Features:
#         - **CD1/CD2 Handling**: **Completely ignores CD1 columns, only uses CD2 values**
#         - **Smart Blank Handling**: If 1+1 CD2 (COMP) is blank but SATP CD2 (TP) has value, uses SATP value and sets Policy=TP
#         - **Policy Code Recognition**: 
#           - 1+5 = New Two Wheeler COMP
#           - 1+1 = Old PVT car and TW COMP (if has value)
#           - 1+3 = New PVT car COMP
#           - SATP = TP (Third Party)
#         - **Segment Types**:
#           - TW: 1+5 (New), 1+1 (Old), includes bikes
#           - PVT CAR: COMP + SAOD (+ means OR), TP
#           - CV: GVW, PCV 3W, GCV 3W
#           - BUS: School Bus, Staff Bus
#           - TAXI: includes PVT Taxi
#           - MISD: Tractor, Cranes, Garbage Vans
        
#         ### 📊 Critical CD2 Logic:
#         - **If CD1 columns exist → IGNORE COMPLETELY**
#         - **Only extract from CD2 columns**
#         - **If 1+1 CD2 blank but SATP CD2 has value → Use SATP, Policy=TP**
#         - **If both have values → Create 2 records (COMP + TP)**
        
#         ### 📈 Output Format:
#         ```
#         Location    | Segment     | Policy | Payin | Payout | Remarks
#         RJ_Good     | TW 1+5      | COMP   | 30%   | 27%    | New Two Wheeler
#         RJ_Good     | TW SATP     | TP     | 25%   | 23%    | Third Party Only
#         ```
#         """)
#         return

#     if process_button:
#         try:
#             # Read file contents
#             policy_file_bytes = policy_file.read()
            
#             # Process files
#             with st.spinner("Processing Digit policy file... This may take a few moments."):
#                 results = process_files(
#                     policy_file_bytes, policy_file.name, policy_file.type,
#                     company_name
#                 )
            
#             st.success("🎉 Digit format processing completed successfully!")
            
#             # Display results in tabs
#             tab1, tab2, tab3, tab4 = st.tabs([
#                 "📊 Final Results", 
#                 "📁 Extracted Text", 
#                 "🧾 Parsed Data", 
#                 "📥 Download"
#             ])
            
#             with tab1:
#                 st.subheader("📊 Processed Digit Policy Data")
                
#                 # Display in the requested format
#                 st.markdown("### Location | Segment | Policy | Payin | Payout | Remarks")
#                 st.dataframe(results["df_calc"], use_container_width=True)
                
#                 # Summary statistics
#                 col1, col2, col3, col4 = st.columns(4)
#                 with col1:
#                     st.metric("Total Records", len(results["calculated_data"]))
#                 with col2:
#                     # Calculate average payin from the processed data
#                     payin_values = []
#                     for record in results["calculated_data"]:
#                         payin_str = record.get('Payin', '0%').replace('%', '')
#                         try:
#                             payin_values.append(float(payin_str))
#                         except:
#                             payin_values.append(0)
#                     avg_payin = sum(payin_values) / len(payin_values) if payin_values else 0
#                     st.metric("Avg Payin", f"{avg_payin:.1f}%")
#                 with col3:
#                     segments = set([r.get('Segment', 'N/A') for r in results["calculated_data"]])
#                     st.metric("Unique Segments", len(segments))
#                 with col4:
#                     st.metric("Company", company_name)
                
#                 # Show sample of formula application
#                 st.subheader("🔧 Sample Calculations")
#                 if results["calculated_data"]:
#                     sample_record = results["calculated_data"][0]
#                     st.write(f"**Example**: {sample_record.get('Segment')} policy with {sample_record.get('Payin')} payin → {sample_record.get('Payout')} payout (Explanation: {sample_record.get('Rule Explanation')})")
                
#                 # Show CD2 extraction info
#                 st.subheader("ℹ️ CD2 Processing Info")
#                 st.info("✅ CD1 columns were ignored. Only CD2 values used for Payin calculation.")
#                 st.info("✅ Blank 1+1 CD2 with SATP CD2 value → Used SATP and set Policy=TP")
            
#             with tab2:
#                 st.subheader("📁 Extracted Text from Digit Policy File")
#                 st.text_area(
#                     "Policy Text", 
#                     results["extracted_text"], 
#                     height=400,
#                     key="policy_text"
#                 )
                
#                 st.subheader("📊 Embedded Formula Rules")
#                 st.write("The system uses these formula rules:")
#                 df_formula = pd.DataFrame(FORMULA_DATA)
#                 st.dataframe(df_formula, use_container_width=True)
            
#             with tab3:
#                 st.subheader("🧾 Parsed Digit Policy Data (JSON)")
#                 st.json(results["parsed_data"])
                
#                 st.subheader("🧮 Calculated Data with Payouts")
#                 st.json(results["calculated_data"])
            
#             with tab4:
#                 st.subheader("📥 Download Results")
                
#                 # Excel download (fixed format)
#                 st.download_button(
#                     label="📊 Download Excel File (.xlsx)",
#                     data=results["excel_data"],
#                     file_name=f"{company_name}_digit_processed_policies.xlsx",
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                 )
                
#                 # JSON download
#                 json_data = json.dumps(results["calculated_data"], indent=2)
#                 st.download_button(
#                     label="📄 Download JSON Data",
#                     data=json_data,
#                     file_name=f"{company_name}_digit_processed_data.json",
#                     mime="application/json"
#                 )
                
#                 # CSV download
#                 csv_data = results["df_calc"].to_csv(index=False)
#                 st.download_button(
#                     label="📋 Download CSV File",
#                     data=csv_data,
#                     file_name=f"{company_name}_digit_processed_policies.csv",
#                     mime="text/csv"
#                 )
                
#                 st.success("✅ Files include Rule Explanation column")
#                 st.info("💡 The Excel file contains Digit processed data: Location | Segment | Policy | Payin | Payout | Remarks")
#                 st.info("🔍 CD1 ignored, CD2 used. Blank COMP with TP value handled correctly.")
                
#         except Exception as e:
#             st.error(f"❌ Error processing Digit files: {str(e)}")
#             st.exception(e)


# if __name__ == "__main__":
#     main()


import streamlit as st
import pandas as pd
from io import BytesIO
import base64
import json
import os
from dotenv import load_dotenv
import logging
import re

# Check if required packages are available
try:
    from openai import OpenAI
except ImportError:
    st.error("OpenAI package not found. Please install it using 'pip install openai'")
    st.stop()

try:
    import PyPDF2
except ImportError:
    st.warning("PyPDF2 not found. PDF text extraction will use OpenAI vision only.")
    PyPDF2 = None

try:
    import pdfplumber
except ImportError:
    st.warning("pdfplumber not found. PDF text extraction will use PyPDF2 or OpenAI vision only.")
    pdfplumber = None

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    get_column_letter = None

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load OpenAI API key
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    st.error("OPENAI_API_KEY environment variable not set. Please set it in your .env file or environment variables.")
    st.stop()

# Initialize OpenAI client
try:
    client = OpenAI(api_key=OPENAI_API_KEY)
except Exception as e:
    st.error(f"Failed to initialize OpenAI client: {str(e)}")
    st.stop()

# Embedded Formula Data 
FORMULA_DATA = [
    {"LOB": "TW", "SEGMENT": "1+5", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "NIL"},
    {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "NIL"},
    {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-2%", "REMARKS": "Payin Below 20%"},
    {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
    {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
    {"LOB": "TW", "SEGMENT": "TW SAOD + COMP", "INSURER": "DIGIT", "PO": "-5%", "REMARKS": "Payin Above 50%"},
    {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Bajaj, Digit, ICICI", "PO": "-2%", "REMARKS": "Payin Below 20%"},
    {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Bajaj, Digit, ICICI", "PO": "-3%", "REMARKS": "Payin Above 20%"},
    {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
    {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
    {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
    {"LOB": "TW", "SEGMENT": "TW TP", "INSURER": "Rest of Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
    {"LOB": "PVT CAR", "SEGMENT": "PVT CAR COMP + SAOD", "INSURER": "All Companies", "PO": "90% of Payin", "REMARKS": "All Fuel"},
    {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Bajaj, Digit, SBI", "PO": "-2%", "REMARKS": "Payin Below 20%"},
    {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Bajaj, Digit, SBI", "PO": "-3%", "REMARKS": "Payin Above 20%"},
    {"LOB": "PVT CAR", "SEGMENT": "PVT CAR TP", "INSURER": "Rest of Companies", "PO": "90% of Payin", "REMARKS": "Zuno -  21"},
    {"LOB": "CV", "SEGMENT": "Upto 2.5 GVW", "INSURER": "Reliance, SBI, Tata", "PO": "-2%", "REMARKS": "NIL"},
    {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
    {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
    {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
    {"LOB": "CV", "SEGMENT": "All GVW & PCV 3W, GCV 3W", "INSURER": "Rest of Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
    {"LOB": "BUS", "SEGMENT": "SCHOOL BUS", "INSURER": "TATA, Reliance, Digit, ICICI", "PO": "Less 2% of Payin", "REMARKS": "NIL"},
    {"LOB": "BUS", "SEGMENT": "SCHOOL BUS", "INSURER": "Rest of Companies", "PO": "88% of Payin", "REMARKS": "NIL"},
    {"LOB": "BUS", "SEGMENT": "STAFF BUS", "INSURER": "All Companies", "PO": "88% of Payin", "REMARKS": "NIL"},
    {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-2%", "REMARKS": "Payin Below 20%"},
    {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-3%", "REMARKS": "Payin 21% to 30%"},
    {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-4%", "REMARKS": "Payin 31% to 50%"},
    {"LOB": "TAXI", "SEGMENT": "TAXI", "INSURER": "All Companies", "PO": "-5%", "REMARKS": "Payin Above 50%"},
    {"LOB": "MISD", "SEGMENT": "Misd, Tractor", "INSURER": "All Companies", "PO": "88% of Payin", "REMARKS": "NIL"}
]


def extract_text_from_pdf_file(pdf_bytes: bytes) -> str:
    """Extract text from PDF using multiple methods"""
    extracted_text = ""
    
    # Method 1: Try pdfplumber first (most accurate)
    if pdfplumber:
        try:
            with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        extracted_text += page_text + "\n"
            if extracted_text.strip():
                logger.info("PDF text extracted using pdfplumber")
                return extracted_text.strip()
        except Exception as e:
            logger.warning(f"pdfplumber extraction failed: {str(e)}")
    
    # Method 2: Try PyPDF2 as fallback
    if PyPDF2:
        try:
            pdf_reader = PyPDF2.PdfReader(BytesIO(pdf_bytes))
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    extracted_text += page_text + "\n"
            if extracted_text.strip():
                logger.info("PDF text extracted using PyPDF2")
                return extracted_text.strip()
        except Exception as e:
            logger.warning(f"PyPDF2 extraction failed: {str(e)}")
    
    # Method 3: Return empty string if all methods fail
    logger.warning("All PDF text extraction methods failed")
    return ""


def extract_text_from_file(file_bytes: bytes, filename: str, content_type: str) -> str:
    """Extract text from uploaded file"""
    file_extension = filename.split('.')[-1].lower() if '.' in filename else ''
    file_type = content_type if content_type else file_extension

    # Image-based extraction with enhanced OCR
    image_extensions = ['png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff']
    if file_extension in image_extensions or file_type.startswith('image/'):
        image_base64 = base64.b64encode(file_bytes).decode('utf-8')
        
        prompt = """Extract all insurance policy text accurately from this image using OCR.
        Pay special attention to identifying:
        - Agency/PB Clusters (these are locations)
        - Agency/PB segments (these are vehicle segments)
        - CD2 values (these are payin percentages - IGNORE CD1 completely)
        - Policy codes like 1+1, 1+3, 1+5, SATP, TP
        - Company names
        - Any percentage values
        - Any numerical data
        - Table structure, including if values are blank under certain columns
        - If CD1 columns exist, ignore them completely, only extract CD2 values
        - If 1+1 CD2 (COMP) is blank but SATP CD2 (TP) has a value, note this relationship
        Extract all text exactly as it appears in the image."""
            
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/{file_extension};base64,{image_base64}"}}
                ]
            }],
            temperature=0.1
        )
        return response.choices[0].message.content.strip()

    # Enhanced PDF extraction
    if file_extension == 'pdf':
        # First try to extract text directly from PDF
        pdf_text = extract_text_from_pdf_file(file_bytes)
        
        if pdf_text and len(pdf_text.strip()) > 50:
            # If we got good text extraction, use it
            logger.info("Using direct PDF text extraction")
            return pdf_text
        else:
            # Fallback to OpenAI vision for PDF
            logger.info("Using OpenAI vision for PDF")
            pdf_base64 = base64.b64encode(file_bytes).decode("utf-8")
            
            prompt = """Extract insurance policy details from this PDF.
            Focus on identifying Agency/PB Clusters (locations), segments, CD2 values (ignore CD1), and policy codes.
            If 1+1 CD2 is blank but SATP CD2 has a value, note this relationship."""
                
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:application/pdf;base64,{pdf_base64}"}}
                    ]
                }],
                temperature=0.1
            )
            return response.choices[0].message.content.strip()

    # Text files
    if file_extension == 'txt':
        return file_bytes.decode('utf-8', errors='ignore')

    # CSV files
    if file_extension == 'csv':
        df = pd.read_csv(BytesIO(file_bytes), encoding='utf-8', errors='ignore')
        return df.to_string()

    # Excel files - Fixed to handle binary Excel files
    if file_extension in ['xlsx', 'xls', 'xlsm']:
        try:
            # Try to read with pandas
            all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine='openpyxl')
            dfs = []
            for sheet_name, df_sheet in all_sheets.items():
                df_sheet["Source_Sheet"] = sheet_name
                dfs.append(df_sheet)
            df = pd.concat(dfs, ignore_index=True, join="outer")
            return df.to_string(index=False)
        except Exception as e:
            # If pandas fails, try with xlrd for older Excel files
            try:
                all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, engine='xlrd')
                dfs = []
                for sheet_name, df_sheet in all_sheets.items():
                    df_sheet["Source_Sheet"] = sheet_name
                    dfs.append(df_sheet)
                df = pd.concat(dfs, ignore_index=True, join="outer")
                return df.to_string(index=False)
            except Exception as e2:
                logger.error(f"Failed to read Excel file with both engines: {str(e)}, {str(e2)}")
                raise ValueError(f"Could not read Excel file: {filename}. Error: {str(e)}")

    raise ValueError(f"Unsupported file type for {filename}")


def clean_json_response(response_text: str) -> str:
    """Clean and extract JSON from OpenAI response"""
    # Remove markdown code blocks
    cleaned = re.sub(r'```json\s*', '', response_text)
    cleaned = re.sub(r'```\s*$', '', cleaned)
    
    # Remove any text before the first [ or {
    json_start = -1
    for i, char in enumerate(cleaned):
        if char in '[{':
            json_start = i
            break
    
    if json_start != -1:
        cleaned = cleaned[json_start:]
    
    # Remove any text after the last ] or }
    json_end = -1
    for i in range(len(cleaned) - 1, -1, -1):
        if cleaned[i] in ']}':
            json_end = i + 1
            break
    
    if json_end != -1:
        cleaned = cleaned[:json_end]
    
    return cleaned.strip()


def ensure_list_format(data) -> list:
    """Ensure data is in list format"""
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        return [data]  # Convert single object to list
    else:
        raise ValueError(f"Expected list or dict, got {type(data)}")


def classify_payin(payin_str):
    """
    Converts Payin string (e.g., '50%') to float and classifies its range.
    """
    try:
        # Handle various formats
        payin_clean = str(payin_str).replace('%', '').replace(' ', '')
        payin_value = float(payin_clean)
        
        if payin_value <= 20:
            category = "Payin Below 20%"
        elif 21 <= payin_value <= 30:
            category = "Payin 21% to 30%"
        elif 31 <= payin_value <= 50:
            category = "Payin 31% to 50%"
        else:
            category = "Payin Above 50%"
        return payin_value, category
    except (ValueError, TypeError):
        logger.warning(f"Could not parse payin value: {payin_str}")
        return 0.0, "Payin Below 20%"


def apply_formula_directly(policy_data, company_name):
    """
    Apply formula rules directly using Python logic based on segment policy type, insurer, and remarks
    """
    calculated_data = []
    
    for record in policy_data:
        # Get policy details
        segment = record.get('Segment', '').upper() or ''
        policy_type = (record.get('Policy', '') or '').upper()
        payin_value = float(record.get('Payin', '0%').replace('%', '').replace(' ', '') or 0)
        location = record.get('Location', 'N/A')
        remarks = record.get('Remarks', 'NIL')

        # Determine LOB from segment
        lob = ""
        if any(tw_keyword in segment for tw_keyword in ['TW', '2W', 'TWO WHEELER', 'BIKE']):
            lob = "TW"
        elif any(car_keyword in segment for car_keyword in ['PVT CAR', 'PRIVATE CAR', 'CAR']):
            lob = "PVT CAR"
        elif any(cv_keyword in segment for cv_keyword in ['CV', 'COMMERCIAL', 'LCV', 'GVW', 'TN', 'PCV', 'GCV']):
            lob = "CV"
        elif 'BUS' in segment:
            lob = "BUS"
        elif any(taxi_keyword in segment for taxi_keyword in ['TAXI', 'PVT TAXI']):
            lob = "TAXI"
        elif any(misd_keyword in segment for misd_keyword in ['MISD', 'TRACTOR', 'MISC', 'CRANE', 'GARBAGE']):
            lob = "MISD"
        else:
            lob = "TW"  # Default to TW if uncertain

        # Extract policy type from segment if present
        segment_policy_type = None
        if 'COMP' in segment or 'SAOD' in segment or 'TP' in segment:
            segment_policy_type = next((pt for pt in ['COMP', 'SAOD', 'TP'] if pt in segment), None)
        else:
            segment_policy_type = policy_type  # Fall back to record's policy type

        # Find matching formula rule based on segment, insurer, and remarks
        matched_rule = None
        rule_explanation = ""
        company_normalized = company_name.upper().replace('GENERAL', '').replace('INSURANCE', '').strip()

        for rule in FORMULA_DATA:
            if rule["LOB"] != lob:
                continue

            # Check segment match including policy type from segment
            rule_segment = rule["SEGMENT"].upper()
            segment_match = False
            if lob == "TW":
                if "1+5" in rule_segment and "1+5" in segment:
                    segment_match = True
                elif "SAOD + COMP" in rule_segment and segment_policy_type in ["COMP", "SAOD"]:
                    segment_match = True
                elif "TP" in rule_segment and segment_policy_type == "TP":
                    segment_match = True
            elif lob == "PVT CAR":
                if "COMP + SAOD" in rule_segment and segment_policy_type in ["COMP", "SAOD"]:
                    segment_match = True
                elif "TP" in rule_segment and segment_policy_type == "TP":
                    segment_match = True
            elif lob == "CV":
                if "UPTO 2.5 GVW" in rule_segment and "UPTO 2.5 GVW" in segment:
                    segment_match = True
                elif "ALL GVW" in rule_segment and any(gvw in segment for gvw in ["GVW", "PCV 3W", "GCV 3W"]):
                    segment_match = True
            elif lob == "BUS":
                if "SCHOOL BUS" in rule_segment and "SCHOOL" in segment:
                    segment_match = True
                elif "STAFF BUS" in rule_segment and "STAFF" in segment:
                    segment_match = True
            elif lob == "TAXI":
                segment_match = True
            elif lob == "MISD":
                segment_match = True

            if not segment_match:
                continue

            # Check insurer match
            insurers = [ins.strip().upper() for ins in rule["INSURER"].split(',')]
            company_match = False
            if "ALL COMPANIES" in insurers:
                company_match = True
            elif "REST OF COMPANIES" in insurers:
                is_in_specific_list = False
                for other_rule in FORMULA_DATA:
                    if (other_rule["LOB"] == rule["LOB"] and 
                        other_rule["SEGMENT"] == rule["SEGMENT"] and
                        "REST OF COMPANIES" not in other_rule["INSURER"] and
                        "ALL COMPANIES" not in other_rule["INSURER"]):
                        other_insurers = [ins.strip().upper() for ins in other_rule["INSURER"].split(',')]
                        if any(company_key in company_normalized for company_key in other_insurers):
                            is_in_specific_list = True
                            break
                if not is_in_specific_list:
                    company_match = True
            else:
                for insurer in insurers:
                    if insurer in company_normalized or company_normalized in insurer:
                        company_match = True
                        break

            if not company_match:
                continue

            # Check remarks match
            rule_remarks = rule.get("REMARKS", "").upper()
            if rule_remarks == "NIL" or "NIL" in rule_remarks:
                matched_rule = rule
                rule_explanation = f"Match: LOB={lob}, Segment={rule_segment}, Insurer={rule['INSURER']}, Remarks=NIL"
                break
            elif any(remark_keyword in rule_remarks for remark_keyword in ["PAYIN BELOW", "PAYIN 21%", "PAYIN 31%", "PAYIN ABOVE"]):
                if any(remark_keyword in remarks.upper() for remark_keyword in ["PAYIN BELOW", "PAYIN 21%", "PAYIN 31%", "PAYIN ABOVE"]):
                    matched_rule = rule
                    rule_explanation = f"Match: LOB={lob}, Segment={rule_segment}, Insurer={rule['INSURER']}, Remarks={rule_remarks}"
                    break
            else:
                matched_rule = rule
                rule_explanation = f"Match: LOB={lob}, Segment={rule_segment}, Insurer={rule['INSURER']}, Remarks={rule_remarks}"
                break

        # Calculate payout based on matched rule
        if matched_rule:
            po_formula = matched_rule["PO"]
            calculated_payout = 0
            
            if "90% of Payin" in po_formula:
                calculated_payout = payin_value * 0.9
            elif "88% of Payin" in po_formula:
                calculated_payout = payin_value * 0.88
            elif "Less 2% of Payin" in po_formula:
                calculated_payout = payin_value - 2
            elif "-2%" in po_formula:
                calculated_payout = payin_value - 2
            elif "-3%" in po_formula:
                calculated_payout = payin_value - 3
            elif "-4%" in po_formula:
                calculated_payout = payin_value - 4
            elif "-5%" in po_formula:
                calculated_payout = payin_value - 5
            else:
                calculated_payout = payin_value  # Default to original payin
            
            # Ensure non-negative result
            calculated_payout = max(0, calculated_payout)
        else:
            # No rule matched - use default 90% of payin
            calculated_payout = payin_value * 0.9
            rule_explanation = f"No match for LOB={lob}, Segment={segment}, Insurer={company_name}, using default 90% of Payin"

        # Create result record
        result_record = {
            'Location': location,
            'Segment': segment,
            'Policy': segment_policy_type if segment_policy_type else policy_type,
            'Payin': f"{int(payin_value)}%",
            'Payout': f"{int(calculated_payout)}%",
            'Remarks': remarks,
            'Rule Explanation': rule_explanation
        }
        
        calculated_data.append(result_record)
    
    return calculated_data

def process_files(policy_file_bytes, policy_filename, policy_content_type, company_name):
    """Main processing function"""
    try:
        st.info("📄 Extracting text from policy file...")
        
        # Extract text with enhanced intelligence
        extracted_text = extract_text_from_file(policy_file_bytes, policy_filename, policy_content_type)

        logger.info(f"Extracted text length: {len(extracted_text)}")

        st.success(f"✅ Text extracted successfully! Policy: {len(extracted_text)} chars")

        # Parse policy data with enhanced segment identification for Digit format
        st.info("🧠 Parsing policy data with AI for Digit format...")
        
        parse_prompt = f"""
        Analyze the following insurance policy text from Digit and extract data according to these specific rules:

        Company Name: {company_name}

        IMPORTANT DIGIT FORMAT RULES:
        - "Agency/PB Clusters" = Location (extract this as Location)
        - "Agency/PB segment" = Segment (extract vehicle type)
        - **CD2 = Payin percentage (COMPLETELY IGNORE CD1 - if CD1 columns exist, ignore them entirely)**
        - **If 1+1 CD2 (COMP payin column) is blank, but SATP CD2 (TP column) has a value, use SATP CD2 value and set Policy="TP"**
        
        SEGMENT TYPES FOR DIGIT:
        - **Two Wheeler (TW)**: 1+5 (New TW), 1+1 (Old PVT car and TW), 1+3 (New PVT car)
        - **PVT CAR**: with policy COMP + SAOD (where + means OR operator), also TP
        - **CV**: GVW, PCV 3 Wheeler, GCV 3W
        - **BUS**: School Bus, Staff Bus
        - **TAXI**: includes PVT Taxi
        - **MISD**: Tractor, Cranes, Garbage Vans
        
        POLICY TYPES:
        1. **COMP** (also known as Package/First Party): for Private Car, Two Wheelers, PCV (auto, bus), GCV, MISD
        2. **SAOD**: comes with Private Car and Two Wheeler
        3. **Third Party (TP)**: comes with Private Car, Two Wheeler, P.C.V, G.C.V & MISD
        4. **1+1 means COMP**, **SATP means TP** for formula evaluation

        CRITICAL CD1/CD2 HANDLING:
        - **If CD1 columns are present, completely ignore them**
        - **Only extract values from CD2 columns**
        - **If 1+1 CD2 (COMP) is blank but SATP CD2 (TP) has value, use that value and set Policy="TP"**
        - **If both have values, create separate records for each**

        Extract into JSON records with these exact fields:
        - "Location": from Agency/PB Clusters field
        - "Segment": from Agency/PB segment field (include vehicle type details)
        - "Policy": determine from codes and available values (1+1=COMP if has value, SATP=TP if has value)
        - "Payin": from CD2 values only (convert to percentage format)
        "Remarks": ALWAYS include additional information including vehicle makes (e.g. Hero/Honda), age info (e.g. Upto 180cc), validity, etc. If no extra info, set to "NIL"

        EXAMPLE OUTPUT FORMAT:
        [
          {{
            "Location": "RJ_Good",
            "Segment": "TW 1+5", 
            "Policy": "COMP",
            "Payin": "30%",
            "Remarks": "New Two Wheeler"
          }},
          {{
            "Location": "RJ_Good",
            "Segment": "TW SATP", 
            "Policy": "TP",
            "Payin": "25%",
            "Remarks": "Third Party Only"
          }}
        ]

        CRITICAL INSTRUCTIONS:
        1. **IGNORE ALL CD1 VALUES COMPLETELY**
        2. **ONLY USE CD2 VALUES FOR PAYIN**
        3. **If 1+1 CD2 blank but SATP CD2 has value → use SATP value, set Policy="TP"**
        4. **If both 1+1 CD2 and SATP CD2 have values → create 2 records (one COMP, one TP)**
        5. **1+1 = COMP, SATP = TP for policy determination**
        6. **PVT Taxi goes under TAXI segment**
        7. **MISD includes Tractor, Cranes, Garbage Vans**

        Text to analyze:
        {extracted_text}
        
        Return ONLY a valid JSON array, no other text.
        """
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Extract insurance policy data as JSON array for Digit format. CRITICAL: Ignore CD1 completely, only use CD2. If 1+1 CD2 blank but SATP CD2 has value, use SATP and set Policy=TP. 1+1=COMP, SATP=TP. Return ONLY valid JSON array."},
                {"role": "user", "content": parse_prompt}
            ],
            temperature=0.1
        )
        
        parsed_json = response.choices[0].message.content.strip()
        logger.info(f"Raw OpenAI response: {parsed_json[:500]}...")
        
        # Clean the JSON response
        cleaned_json = clean_json_response(parsed_json)
        logger.info(f"Cleaned JSON: {cleaned_json[:500]}...")
        
        try:
            policy_data = json.loads(cleaned_json)
            policy_data = ensure_list_format(policy_data)
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {str(e)}")
            logger.error(f"Problematic JSON: {cleaned_json}")
            
            # Fallback: create a basic structure
            policy_data = [{
                "Location": "Unknown",
                "Segment": "TW 1+5",
                "Policy": "COMP", 
                "Payin": "30%",
                "Remarks": f"Error parsing data: {str(e)}"
            }]

        st.success(f"✅ Successfully parsed {len(policy_data)} policy records")

        # Pre-classify Payin values
        for record in policy_data:
            try:
                payin_val, payin_cat = classify_payin(record.get('Payin', '0%'))
                record['Payin_Value'] = payin_val
                record['Payin_Category'] = payin_cat
            except Exception as e:
                logger.warning(f"Error classifying payin for record: {record}, error: {str(e)}")
                record['Payin_Value'] = 30.0  # Default value
                record['Payin_Category'] = "Payin 21% to 30%"

        # Apply formulas directly using Python logic
        st.info("🧮 Applying formulas and calculating payouts...")
        
        calculated_data = apply_formula_directly(policy_data, company_name)

        st.success(f"✅ Successfully calculated {len(calculated_data)} records with payouts")

        # Create Excel file with proper format
        st.info("📊 Creating Excel file...")
        
        df_calc = pd.DataFrame(calculated_data)
        output = BytesIO()
        
        # Use openpyxl engine to avoid binary format issues
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write data starting from row 3 to leave space for headers
            df_calc.to_excel(writer, sheet_name='Policy Data', startrow=2, index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Policy Data']
            
            # number of columns (ensure at least 1 to avoid merge errors)
            cols_count = max(1, len(df_calc.columns))

            # Set company name in row 1 (merged across columns)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols_count)
            company_cell = worksheet.cell(row=1, column=1, value=company_name)
            try:
                company_cell.font = company_cell.font.copy(bold=True, size=14)
                company_cell.alignment = company_cell.alignment.copy(horizontal='center')
            except Exception:
                pass
            
            # Set subtitle in row 2 (merged across columns)
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols_count)
            subtitle_cell = worksheet.cell(row=2, column=1, value='Policy Data with Payin and Calculated Payouts')
            try:
                subtitle_cell.font = subtitle_cell.font.copy(bold=True, size=12)
                subtitle_cell.alignment = subtitle_cell.alignment.copy(horizontal='center')
            except Exception:
                pass
            
            # Apply header formatting to column headers (row 3)
            for col_num, value in enumerate(df_calc.columns.values, 1):
                header_cell = worksheet.cell(row=3, column=col_num, value=value)
                try:
                    header_cell.font = header_cell.font.copy(bold=True)
                except Exception:
                    pass
            
            # Auto-adjust column widths
            for col_idx, column_name in enumerate(df_calc.columns, 1):
                max_length = 0
                # Use get_column_letter (safe) instead of accessing possibly-merged cell.column_letter
                if get_column_letter:
                    column_letter = get_column_letter(col_idx)
                else:
                    # fallback: use simple letter mapping for first 26 columns
                    column_letter = chr(64 + col_idx) if col_idx <= 26 else str(col_idx)

                # Check header length
                max_length = max(max_length, len(str(column_name)))
                
                # Check data length
                for row_idx in range(4, len(df_calc) + 4):  # Start from row 4 (data starts there)
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                try:
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                except Exception:
                    # If column_dimensions fails for any reason, ignore and continue
                    pass

        output.seek(0)
        excel_data = output.read()

        return {
            "extracted_text": extracted_text,
            "parsed_data": policy_data,
            "calculated_data": calculated_data,
            "excel_data": excel_data,
            "df_calc": df_calc
        }

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}", exc_info=True)
        raise Exception(f"An error occurred: {str(e)}")


# Streamlit App
def main():
    st.set_page_config(
        page_title="Insurance Policy Processing - Digit Format", 
        page_icon="📋", 
        layout="wide"
    )
    
    st.title("🏢 Insurance Policy Processing System - Digit Format")
    st.markdown("---")
    
    # Sidebar for inputs
    with st.sidebar:
        st.header("📁 File Upload")
        
        # Company name input
        company_name = st.text_input(
            "Company Name", 
            value="Digit",
            help="Enter the insurance company name"
        )
        
        # Policy file upload
        policy_file = st.file_uploader(
            "📄 Upload Policy File",
            type=['pdf', 'png', 'jpg', 'jpeg', 'txt', 'csv', 'xlsx', 'xls', 'xlsm'],
            help="Upload your insurance policy document from Digit"
        )
        
        # Show Digit format mapping info
        st.info("""
        📊 **Digit Format Mapping:**
        - Agency/PB Clusters → Location
        - Agency/PB segment → Segment  
        - **CD2 → Payin (CD1 IGNORED)**
        - **1+1 CD2 blank + SATP CD2 value → Use SATP, Policy=TP**
        - 1+5 → New TW COMP
        - 1+1 → Old COMP (if has value)
        - 1+3 → New PVT COMP
        - SATP → TP
        - PVT Taxi → TAXI
        - MISD includes Tractor, Cranes, Garbage Vans
        """)
        
        # Process button
        process_button = st.button(
            "🚀 Process Digit Policy File", 
            type="primary",
            disabled=not policy_file
        )
    
    # Main content area
    if not policy_file:
        st.info("👆 Please upload a Digit policy file to begin processing.")
        st.markdown("""
        ### 📋 Digit Format Instructions:
        1. **Company Name**: Enter the insurance company name (default: Digit)
        2. **Policy File**: Upload the Digit document containing policy data
        3. **Process**: Click the process button to extract data and calculate payouts
        
        ### 🎯 Digit Format Key Features:
        - **CD1/CD2 Handling**: **Completely ignores CD1 columns, only uses CD2 values**
        - **Smart Blank Handling**: If 1+1 CD2 (COMP) is blank but SATP CD2 (TP) has value, uses SATP value and sets Policy=TP
        - **Policy Code Recognition**: 
          - 1+5 = New Two Wheeler COMP
          - 1+1 = Old PVT car and TW COMP (if has value)
          - 1+3 = New PVT car COMP
          - SATP = TP (Third Party)
        - **Segment Types**:
          - TW: 1+5 (New), 1+1 (Old), includes bikes
          - PVT CAR: COMP + SAOD (+ means OR), TP
          - CV: GVW, PCV 3W, GCV 3W
          - BUS: School Bus, Staff Bus
          - TAXI: includes PVT Taxi
          - MISD: Tractor, Cranes, Garbage Vans
        
        ### 📊 Critical CD2 Logic:
        - **If CD1 columns exist → IGNORE COMPLETELY**
        - **Only extract from CD2 columns**
        - **If 1+1 CD2 blank but SATP CD2 has value → Use SATP, Policy=TP**
        - **If both have values → Create 2 records (COMP + TP)**
        
        ### 📈 Output Format:
        ```
        Location    | Segment     | Policy | Payin | Payout | Remarks
        RJ_Good     | TW 1+5      | COMP   | 30%   | 27%    | New Two Wheeler
        RJ_Good     | TW SATP     | TP     | 25%   | 23%    | Third Party Only
        ```
        """)
        return

    if process_button:
        try:
            # Read file contents
            policy_file_bytes = policy_file.read()
            
            # Process files
            with st.spinner("Processing Digit policy file... This may take a few moments."):
                results = process_files(
                    policy_file_bytes, policy_file.name, policy_file.type,
                    company_name
                )
            
            st.success("🎉 Digit format processing completed successfully!")
            
            # Display results in tabs
            tab1, tab2, tab3, tab4 = st.tabs([
                "📊 Final Results", 
                "📁 Extracted Text", 
                "🧾 Parsed Data", 
                "📥 Download"
            ])
            
            with tab1:
                st.subheader("📊 Processed Digit Policy Data")
                
                # Display in the requested format
                st.markdown("### Location | Segment | Policy | Payin | Payout | Remarks")
                st.dataframe(results["df_calc"], use_container_width=True)
                
                # Summary statistics
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Records", len(results["calculated_data"]))
                with col2:
                    # Calculate average payin from the processed data
                    payin_values = []
                    for record in results["calculated_data"]:
                        payin_str = record.get('Payin', '0%').replace('%', '')
                        try:
                            payin_values.append(float(payin_str))
                        except:
                            payin_values.append(0)
                    avg_payin = sum(payin_values) / len(payin_values) if payin_values else 0
                    st.metric("Avg Payin", f"{avg_payin:.1f}%")
                with col3:
                    segments = set([r.get('Segment', 'N/A') for r in results["calculated_data"]])
                    st.metric("Unique Segments", len(segments))
                with col4:
                    st.metric("Company", company_name)
                
                # Show sample of formula application
                st.subheader("🔧 Sample Calculations")
                if results["calculated_data"]:
                    sample_record = results["calculated_data"][0]
                    st.write(f"**Example**: {sample_record.get('Segment')} policy with {sample_record.get('Payin')} payin → {sample_record.get('Payout')} payout (Explanation: {sample_record.get('Rule Explanation')})")
                
                # Show CD2 extraction info
                st.subheader("ℹ️ CD2 Processing Info")
                st.info("✅ CD1 columns were ignored. Only CD2 values used for Payin calculation.")
                st.info("✅ Blank 1+1 CD2 with SATP CD2 value → Used SATP and set Policy=TP")
            
            with tab2:
                st.subheader("📁 Extracted Text from Digit Policy File")
                st.text_area(
                    "Policy Text", 
                    results["extracted_text"], 
                    height=400,
                    key="policy_text"
                )
                
                st.subheader("📊 Embedded Formula Rules")
                st.write("The system uses these formula rules:")
                df_formula = pd.DataFrame(FORMULA_DATA)
                st.dataframe(df_formula, use_container_width=True)
            
            with tab3:
                st.subheader("🧾 Parsed Digit Policy Data (JSON)")
                st.json(results["parsed_data"])
                
                st.subheader("🧮 Calculated Data with Payouts")
                st.json(results["calculated_data"])
            
            with tab4:
                st.subheader("📥 Download Results")
                
                # Excel download (fixed format)
                st.download_button(
                    label="📊 Download Excel File (.xlsx)",
                    data=results["excel_data"],
                    file_name=f"{company_name}_digit_processed_policies.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # JSON download
                json_data = json.dumps(results["calculated_data"], indent=2)
                st.download_button(
                    label="📄 Download JSON Data",
                    data=json_data,
                    file_name=f"{company_name}_digit_processed_data.json",
                    mime="application/json"
                )
                
                # CSV download
                csv_data = results["df_calc"].to_csv(index=False)
                st.download_button(
                    label="📋 Download CSV File",
                    data=csv_data,
                    file_name=f"{company_name}_digit_processed_policies.csv",
                    mime="text/csv"
                )
                
                st.success("✅ Files include Rule Explanation column")
                st.info("💡 The Excel file contains Digit processed data: Location | Segment | Policy | Payin | Payout | Remarks")
                st.info("🔍 CD1 ignored, CD2 used. Blank COMP with TP value handled correctly.")
                
        except Exception as e:
            st.error(f"❌ Error processing Digit files: {str(e)}")
            st.exception(e)


if __name__ == "__main__":
    main()
