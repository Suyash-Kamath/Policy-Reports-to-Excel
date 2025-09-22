# # import os
# # import re
# # import json
# # import base64
# # import pandas as pd
# # from io import BytesIO
# # import streamlit as st
# # from dotenv import load_dotenv
# # from openai import OpenAI

# # # Load API Key
# # load_dotenv()
# # OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# # if not OPENAI_API_KEY:
# #     st.error("⚠️ OPENAI_API_KEY not set in .env or environment variables.")
# #     st.stop()

# # client = OpenAI(api_key=OPENAI_API_KEY)

# # # ---- Prompt Template ----
# # PROMPT_TEMPLATE = """
# # You are given insurance policy text. 
# # Extract and normalize into JSON with these fields:
# # - Segment
# # - Location
# # - Policy Type
# # - Doable Districts
# # - Pay in
# # - Remarks

# # Rules:
# # - If Location is missing but RTOs are mentioned, set Location = "RTOs mentioned".
# # - Doable Districts means the RTOs or places mentioned.
# # - If Policy Type is not explicitly mentioned in the text, set it as "COMP/TP".
# # - Keep ALL extra info (vehicle make, age, validity, etc.) in Remarks.
# # Return only a JSON array.
# # Text:
# # {text}
# # """

# # def clean_json_response(response_text: str) -> str:
# #     """Remove markdown formatting and keep pure JSON"""
# #     cleaned = re.sub(r"```json|```", "", response_text).strip()
# #     return cleaned

# # def extract_text(file_bytes: bytes, filename: str) -> str:
# #     """Extracts text from different file formats"""
# #     ext = filename.split(".")[-1].lower()

# #     if ext == "txt":
# #         return file_bytes.decode("utf-8", errors="ignore")

# #     if ext == "csv":
# #         df = pd.read_csv(BytesIO(file_bytes))
# #         return df.to_string()

# #     if ext in ["xlsx", "xls"]:
# #         sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None)
# #         return "\n".join([df.to_string() for _, df in sheets.items()])

# #     if ext == "pdf":
# #         import pdfplumber
# #         text = ""
# #         with pdfplumber.open(BytesIO(file_bytes)) as pdf:
# #             for page in pdf.pages:
# #                 t = page.extract_text()
# #                 if t:
# #                     text += t + "\n"
# #         return text

# #     if ext in ["jpg", "jpeg", "png"]:
# #         b64 = base64.b64encode(file_bytes).decode("utf-8")
# #         resp = client.chat.completions.create(
# #             model="gpt-4o",
# #             messages=[{"role": "user", "content": [
# #                 {"type": "text", "text": "Extract insurance policy text accurately"},
# #                 {"type": "image_url", "image_url": {"url": f"data:image/{ext};base64,{b64}"}}
# #             ]}],
# #             temperature=0
# #         )
# #         return resp.choices[0].message.content

# #     return ""

# # def process_file(file_bytes: bytes, filename: str):
# #     """Process the uploaded file and return a dataframe"""
# #     text = extract_text(file_bytes, filename)
# #     prompt = PROMPT_TEMPLATE.format(text=text[:6000])  # truncate if too long

# #     resp = client.chat.completions.create(
# #         model="gpt-4o-mini",
# #         messages=[{"role": "user", "content": prompt}],
# #         temperature=0
# #     )
# #     result = clean_json_response(resp.choices[0].message.content)

# #     try:
# #         records = json.loads(result)
# #     except:
# #         records = [{
# #             "Segment":"N/A","Location":"N/A","Policy Type":"COMP/TP",
# #             "Doable Districts":"N/A","Pay in":"N/A","Remarks":result
# #         }]

# #     # ✅ Ensure Policy Type default
# #     for rec in records:
# #         if not rec.get("Policy Type") or rec["Policy Type"].strip() in ["", "N/A"]:
# #             rec["Policy Type"] = "COMP/TP"

# #     df = pd.DataFrame(records)
# #     return df

# # # ---------------- Streamlit UI ---------------- #
# # st.set_page_config(page_title="Insurance Report Normalizer", page_icon="📊", layout="wide")
# # st.title("📊 Insurance Report Normalizer")
# # st.write("Upload your insurance report (PDF, Excel, CSV, TXT, Image) to convert into a structured Excel sheet.")

# # uploaded_file = st.file_uploader(
# #     "Upload Report File",
# #     type=["pdf", "xlsx", "xls", "csv", "txt", "jpg", "jpeg", "png"]
# # )

# # if uploaded_file:
# #     st.info(f"Processing file: {uploaded_file.name} ...")
# #     try:
# #         df = process_file(uploaded_file.read(), uploaded_file.name)
# #         st.success("✅ File processed successfully!")
# #         st.subheader("Preview of Normalized Data")
# #         st.dataframe(df, use_container_width=True)

# #         # Download buttons
# #         excel_output = BytesIO()
# #         df.to_excel(excel_output, index=False)
# #         excel_output.seek(0)

# #         csv_output = df.to_csv(index=False).encode("utf-8")
# #         json_output = df.to_json(orient="records", indent=2).encode("utf-8")

# #         st.download_button("⬇️ Download Excel", excel_output,
# #                            file_name="normalized_output.xlsx",
# #                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
# #         st.download_button("⬇️ Download CSV", csv_output,
# #                            file_name="normalized_output.csv", mime="text/csv")
# #         st.download_button("⬇️ Download JSON", json_output,
# #                            file_name="normalized_output.json", mime="application/json")

# #     except Exception as e:
# #         st.error(f"❌ Error: {e}")


# import os
# import re
# import json
# import base64
# import pandas as pd
# from io import BytesIO
# import streamlit as st
# from dotenv import load_dotenv
# from openai import OpenAI

# # Load API Key
# load_dotenv()
# OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# if not OPENAI_API_KEY:
#     st.error("⚠️ OPENAI_API_KEY not set in .env or environment variables.")
#     st.stop()

# client = OpenAI(api_key=OPENAI_API_KEY)

# # ---- Prompt Template ----
# PROMPT_TEMPLATE = """
# You are given insurance policy text. 
# Extract and normalize into JSON with these fields:
# - Segment
# - Location
# - Policy Type
# - Doable Districts
# - Pay in
# - Remarks

# Rules:
# - If Location is missing but RTOs are mentioned, set Location = "RTOs mentioned".
# - Doable Districts means the RTOs or places mentioned.
# - If Policy Type is not explicitly mentioned in the text, set it as "COMP/TP".
# - Keep ALL extra info (vehicle make, age, validity, etc.) in Remarks.
# Return only a JSON array.
# Text:
# {text}
# """

# def clean_json_response(response_text: str) -> str:
#     """Remove markdown formatting and keep pure JSON"""
#     cleaned = re.sub(r"```json|```", "", response_text).strip()
#     return cleaned

# def extract_text(file_bytes: bytes, filename: str) -> str:
#     """Extracts text from different file formats"""
#     ext = filename.split(".")[-1].lower()

#     if ext == "txt":
#         return file_bytes.decode("utf-8", errors="ignore")

#     if ext == "csv":
#         df = pd.read_csv(BytesIO(file_bytes))
#         return df.to_string()

#     if ext in ["xlsx", "xls"]:
#         sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None)
#         return "\n".join([df.to_string() for _, df in sheets.items()])

#     if ext == "pdf":
#         import pdfplumber
#         text = ""
#         with pdfplumber.open(BytesIO(file_bytes)) as pdf:
#             for page in pdf.pages:
#                 t = page.extract_text()
#                 if t:
#                     text += t + "\n"
#         return text

#     if ext in ["jpg", "jpeg", "png"]:
#         b64 = base64.b64encode(file_bytes).decode("utf-8")
#         resp = client.chat.completions.create(
#             model="gpt-4o",
#             messages=[{"role": "user", "content": [
#                 {"type": "text", "text": "Extract insurance policy text accurately"},
#                 {"type": "image_url", "image_url": {"url": f"data:image/{ext};base64,{b64}"}}
#             ]}],
#             temperature=0.1
#         )
#         return resp.choices[0].message.content

#     return ""

# def process_file(file_bytes: bytes, filename: str, company_name: str):
#     """Process the uploaded file and return a dataframe"""
#     text = extract_text(file_bytes, filename)
#     prompt = PROMPT_TEMPLATE.format(text=text[:6000])  # truncate if too long

#     resp = client.chat.completions.create(
#         model="gpt-4o-mini",
#         messages=[{"role": "user", "content": prompt}],
#         temperature=0.1
#     )
#     result = clean_json_response(resp.choices[0].message.content)

#     try:
#         records = json.loads(result)
#     except:
#         records = [{
#             "Segment":"N/A","Location":"N/A","Policy Type":"COMP/TP",
#             "Doable Districts":"N/A","Pay in":"N/A","Remarks":result
#         }]

#     # ✅ Ensure Policy Type default
#     for rec in records:
#         if not rec.get("Policy Type") or rec["Policy Type"].strip() in ["", "N/A"]:
#             rec["Policy Type"] = "COMP/TP"
#         # ✅ Add Insurer/Company Name to each record
#         rec["Insurer"] = company_name

#     df = pd.DataFrame(records)
#     return df

# def make_excel(df: pd.DataFrame, company_name: str) -> bytes:
#     """Create Excel file with company name as header"""
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, sheet_name="Normalized Data", startrow=2, index=False)
#         ws = writer.sheets["Normalized Data"]

#         # Headers
#         headers = list(df.columns)
#         for col_num, value in enumerate(headers, 1):
#             cell = ws.cell(row=3, column=col_num, value=value)
#             cell.font = cell.font.copy(bold=True)

#         # Company name at top row
#         company_cell = ws.cell(row=1, column=1, value=company_name)
#         ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
#         company_cell.font = company_cell.font.copy(bold=True, size=14)
#         company_cell.alignment = company_cell.alignment.copy(horizontal="center")

#         # Title row
#         title_cell = ws.cell(row=2, column=1, value="Normalized Insurance Policy Data")
#         ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
#         title_cell.font = title_cell.font.copy(bold=True, size=12)
#         title_cell.alignment = title_cell.alignment.copy(horizontal="center")

#     output.seek(0)
#     return output.read()

# # ---------------- Streamlit UI ---------------- #
# st.set_page_config(page_title="Insurance Report Normalizer", page_icon="📊", layout="wide")
# st.title("📊 Insurance Report Normalizer")
# st.write("Upload your insurance report (PDF, Excel, CSV, TXT, Image) to convert into a structured Excel sheet.")

# # Sidebar inputs
# with st.sidebar:
#     st.header("Settings")
#     company_name = st.text_input("🏢 Company / Insurer Name", value="Unknown Company")
#     uploaded_file = st.file_uploader(
#         "📂 Upload Report File",
#         type=["pdf", "xlsx", "xls", "csv", "txt", "jpg", "jpeg", "png"]
#     )

# if uploaded_file and company_name:
#     st.info(f"Processing file: {uploaded_file.name} ...")
#     try:
#         df = process_file(uploaded_file.read(), uploaded_file.name, company_name)
#         st.success("✅ File processed successfully!")
#         st.subheader("Preview of Normalized Data")
#         st.dataframe(df, use_container_width=True)

#         # Export files
#         excel_data = make_excel(df, company_name)
#         csv_data = df.to_csv(index=False).encode("utf-8")
#         json_data = df.to_json(orient="records", indent=2).encode("utf-8")

#         st.subheader("📥 Download Results")
#         st.download_button("⬇️ Download Excel", excel_data,
#                            file_name=f"{company_name}_normalized.xlsx",
#                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#         st.download_button("⬇️ Download CSV", csv_data,
#                            file_name=f"{company_name}_normalized.csv", mime="text/csv")
#         st.download_button("⬇️ Download JSON", json_data,
#                            file_name=f"{company_name}_normalized.json", mime="application/json")

#     except Exception as e:
#         st.error(f"❌ Error: {e}")


# import os
# import re
# import json
# import base64
# import pandas as pd
# from io import BytesIO
# import streamlit as st
# from dotenv import load_dotenv
# from openai import OpenAI

# # Load API Key
# load_dotenv()
# OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# if not OPENAI_API_KEY:
#     st.error("⚠️ OPENAI_API_KEY not set in .env or environment variables.")
#     st.stop()

# client = OpenAI(api_key=OPENAI_API_KEY)

# # ---- Prompt Template ----
# PROMPT_TEMPLATE = """
# You are given insurance policy text or tabular data.
# Extract and normalize into JSON with these fields:
# - Segment
# - Location
# - Policy Type
# - Doable Districts
# - Pay in
# - Remarks
# - Insurer

# Rules:
# - If Location is present, put it in "Location".
# - If RTO state is mentioned, add it into "Remarks" as "RTO State: <state>".
# - Doable Districts means the RTOs or places mentioned.
# - Policy Type may appear as "Types", "Type Comp/TP", "AOTP", "All" — keep it EXACTLY as written in the data.
# - If Policy Type is missing, set it as "COMP/TP".
# - Preserve duplicates: if multiple fields with same name exist, capture all.
# - Headers may sometimes be in rows or sideways — detect intelligently.
# - Keep ALL extra info (vehicle make, age, validity, etc.) in Remarks.

# Return only a JSON array.
# Text:
# {text}
# """

# def clean_json_response(response_text: str) -> str:
#     """Remove markdown formatting and keep pure JSON"""
#     cleaned = re.sub(r"```json|```", "", response_text).strip()
#     return cleaned

# def extract_excel(file_bytes: bytes) -> str:
#     """Extract text from Excel with duplicate headers & flexible header detection"""
#     # Read all sheets, keep duplicates
#     xls = pd.ExcelFile(BytesIO(file_bytes))
#     all_text = []

#     for sheet in xls.sheet_names:
#         try:
#             # Try with header row first
#             df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=0, mangle_dupe_cols=False)
#         except:
#             # Fallback if headers are messy
#             df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=None, mangle_dupe_cols=False)

#         # Detect header row by looking for known keywords
#         for i in range(min(5, len(df))):
#             row_vals = [str(v).lower() for v in df.iloc[i].values if pd.notna(v)]
#             if any(h in " ".join(row_vals) for h in ["segment", "type", "policy", "rto", "doable"]):
#                 df.columns = df.iloc[i]  # set this row as header
#                 df = df.drop(i)
#                 break

#         df["Source_Sheet"] = sheet
#         all_text.append(df.to_string())

#     return "\n".join(all_text)

# def extract_text(file_bytes: bytes, filename: str) -> str:
#     """Extracts text from different file formats"""
#     ext = filename.split(".")[-1].lower()

#     if ext == "txt":
#         return file_bytes.decode("utf-8", errors="ignore")

#     if ext == "csv":
#         df = pd.read_csv(BytesIO(file_bytes), mangle_dupe_cols=False)
#         return df.to_string()

#     if ext in ["xlsx", "xls"]:
#         return extract_excel(file_bytes)

#     if ext == "pdf":
#         import pdfplumber
#         text = ""
#         with pdfplumber.open(BytesIO(file_bytes)) as pdf:
#             for page in pdf.pages:
#                 t = page.extract_text()
#                 if t:
#                     text += t + "\n"
#         return text

#     if ext in ["jpg", "jpeg", "png"]:
#         b64 = base64.b64encode(file_bytes).decode("utf-8")
#         resp = client.chat.completions.create(
#             model="gpt-4o",
#             messages=[{"role": "user", "content": [
#                 {"type": "text", "text": "Extract insurance policy text accurately"},
#                 {"type": "image_url", "image_url": {"url": f"data:image/{ext};base64,{b64}"}}
#             ]}],
#             temperature=0.1
#         )
#         return resp.choices[0].message.content

#     return ""

# def process_file(file_bytes: bytes, filename: str, company_name: str):
#     """Process the uploaded file and return a dataframe"""
#     text = extract_text(file_bytes, filename)
#     prompt = PROMPT_TEMPLATE.format(text=text[:12000])  

#     resp = client.chat.completions.create(
#         model="gpt-4o-mini",
#         messages=[{"role": "user", "content": prompt}],
#         temperature=0.1
#     )
#     result = clean_json_response(resp.choices[0].message.content)

#     try:
#         records = json.loads(result)
#     except:
#         records = [{
#             "Segment":"N/A","Location":"N/A","Policy Type":"COMP/TP",
#             "Doable Districts":"N/A","Pay in":"N/A","Remarks":result
#         }]

#     # ✅ Apply insurer and refine Policy Type handling
#     for rec in records:
#         # Add insurer
#         rec["Insurer"] = company_name

#         # Policy Type rules
#         pt = rec.get("Policy Type", "").strip()
#         if not pt or pt in ["", "N/A"]:
#             rec["Policy Type"] = "COMP/TP"
#         else:
#             # Preserve exactly what's written, e.g. AOTP, All, Comp
#             rec["Policy Type"] = pt  

#     df = pd.DataFrame(records)
#     return df

# def make_excel(df: pd.DataFrame, company_name: str) -> bytes:
#     """Create Excel file with company name as header"""
#     from openpyxl import Workbook
#     from openpyxl.styles import Alignment, Font

#     output = BytesIO()
#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, sheet_name="Normalized Data", startrow=2, index=False)
#         ws = writer.sheets["Normalized Data"]

#         headers = list(df.columns)
#         for col_num, value in enumerate(headers, 1):
#             cell = ws.cell(row=3, column=col_num, value=value)
#             cell.font = Font(bold=True)

#         # Company name top row
#         ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
#         ws.cell(row=1, column=1, value=company_name).font = Font(bold=True, size=14)
#         ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

#         # Title row
#         ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
#         ws.cell(row=2, column=1, value="Normalized Insurance Policy Data").font = Font(bold=True, size=12)
#         ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

#     output.seek(0)
#     return output.read()

# # ---------------- Streamlit UI ---------------- #
# st.set_page_config(page_title="Insurance Report Normalizer", page_icon="📊", layout="wide")
# st.title("📊 Insurance Report Normalizer")

# with st.sidebar:
#     st.header("Settings")
#     company_name = st.text_input("🏢 Company / Insurer Name", value="Unknown Company")
#     uploaded_file = st.file_uploader(
#         "📂 Upload Report File",
#         type=["pdf", "xlsx", "xls", "csv", "txt", "jpg", "jpeg", "png"]
#     )

# if uploaded_file and company_name:
#     st.info(f"Processing file: {uploaded_file.name} ...")
#     try:
#         df = process_file(uploaded_file.read(), uploaded_file.name, company_name)
#         st.success("✅ File processed successfully!")
#         st.subheader("Preview of Normalized Data")
#         st.dataframe(df, use_container_width=True)

#         # Exports
#         excel_data = make_excel(df, company_name)
#         csv_data = df.to_csv(index=False).encode("utf-8")
#         json_data = df.to_json(orient="records", indent=2).encode("utf-8")

#         st.subheader("📥 Download Results")
#         st.download_button("⬇️ Download Excel", excel_data,
#                            file_name=f"{company_name}_normalized.xlsx",
#                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#         st.download_button("⬇️ Download CSV", csv_data,
#                            file_name=f"{company_name}_normalized.csv", mime="text/csv")
#         st.download_button("⬇️ Download JSON", json_data,
#                            file_name=f"{company_name}_normalized.json", mime="application/json")

#     except Exception as e:
#         st.error(f"❌ Error: {e}")




# import os
# import re
# import json
# import base64
# import pandas as pd
# from io import BytesIO
# import streamlit as st
# from dotenv import load_dotenv
# from openai import OpenAI

# # Load API Key
# load_dotenv()
# OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# if not OPENAI_API_KEY:
#     st.error("⚠️ OPENAI_API_KEY not set in .env or environment variables.")
#     st.stop()

# client = OpenAI(api_key=OPENAI_API_KEY)

# # ---- Helper Functions ----
# def clean_json_response(response_text: str) -> str:
#     return re.sub(r"```json|```", "", response_text).strip()

# def parse_transposed_excel(df: pd.DataFrame, company_name: str) -> pd.DataFrame:
#     """Handle reports where first column contains field names, and each other column is a record."""
#     df = df.dropna(how="all").reset_index(drop=True)
#     fields = df.iloc[:, 0].astype(str).str.strip().tolist()

#     records = []
#     for col in df.columns[1:]:
#         values = df[col].tolist()
#         record = dict(zip(fields, values))

#         rec = {
#             "Segment": str(record.get("Segment", "")).strip() or "N/A",
#             "Location": str(record.get("RTO state", "")).strip() or "N/A",
#             "Policy Type": str(record.get("Type Comp/TP", record.get("Type", ""))).strip() or "COMP/TP",
#             "Doable Districts": str(record.get("Doable District Name", "")).strip() or "N/A",
#             "Pay in": str(record.get("Revised PO", "")).strip() or "N/A",
#             "Transaction": str(record.get("Transaction New/Old", "")).strip() or "N/A",
#             "Age": str(record.get("Age", "")).strip() or "N/A",
#             "Remarks": str(record.get("Any Special remarks", "")).strip() or "N/A",
#             "Insurer": company_name
#         }
#         records.append(rec)

#     return pd.DataFrame(records)

# def extract_excel(file_bytes: bytes, company_name: str) -> pd.DataFrame:
#     """Extract and normalize Excel (supports transposed layout and duplicates)."""
#     xls = pd.ExcelFile(BytesIO(file_bytes))
#     all_records = []

#     for sheet in xls.sheet_names:
#         df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=None, mangle_dupe_cols=False)

#         first_col = df.iloc[:, 0].astype(str).str.lower().tolist()
#         if any("segment" in v for v in first_col):
#             parsed = parse_transposed_excel(df, company_name)
#             parsed["Source_Sheet"] = sheet
#             all_records.append(parsed)
#         else:
#             df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, mangle_dupe_cols=False)
#             df["Insurer"] = company_name
#             df["Source_Sheet"] = sheet
#             all_records.append(df)

#     return pd.concat(all_records, ignore_index=True)

# def extract_text(file_bytes: bytes, filename: str, company_name: str):
#     ext = filename.split(".")[-1].lower()

#     if ext in ["xlsx", "xls"]:
#         return extract_excel(file_bytes, company_name)

#     if ext == "csv":
#         df = pd.read_csv(BytesIO(file_bytes), mangle_dupe_cols=False)
#         df["Insurer"] = company_name
#         return df

#     # For txt/pdf/images fallback to AI parsing
#     return None

# def make_excel(df: pd.DataFrame, company_name: str) -> bytes:
#     from openpyxl.styles import Alignment, Font
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, sheet_name="Normalized Data", startrow=2, index=False)
#         ws = writer.sheets["Normalized Data"]

#         headers = list(df.columns)
#         for col_num, value in enumerate(headers, 1):
#             cell = ws.cell(row=3, column=col_num, value=value)
#             cell.font = Font(bold=True)

#         ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
#         ws.cell(row=1, column=1, value=company_name).font = Font(bold=True, size=14)
#         ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

#         ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
#         ws.cell(row=2, column=1, value="Normalized Insurance Policy Data").font = Font(bold=True, size=12)
#         ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

#     output.seek(0)
#     return output.read()

# # ---- Streamlit UI ----
# st.set_page_config(page_title="Insurance Report Normalizer", page_icon="📊", layout="wide")
# st.title("📊 Insurance Report Normalizer")

# with st.sidebar:
#     st.header("Settings")
#     company_name = st.text_input("🏢 Company / Insurer Name", value="Unknown Company")
#     uploaded_file = st.file_uploader(
#         "📂 Upload Report File",
#         type=["pdf", "xlsx", "xls", "csv", "txt", "jpg", "jpeg", "png"]
#     )

# if uploaded_file and company_name:
#     st.info(f"Processing file: {uploaded_file.name} ...")
#     try:
#         df = extract_text(uploaded_file.read(), uploaded_file.name, company_name)

#         if df is None:
#             st.error("⚠️ Non-Excel/CSV reports still use AI parsing — not wired here.")
#         else:
#             st.success("✅ File processed successfully!")
#             st.subheader("Preview of Normalized Data")
#             st.dataframe(df, use_container_width=True)

#             excel_data = make_excel(df, company_name)
#             csv_data = df.to_csv(index=False).encode("utf-8")
#             json_data = df.to_json(orient="records", indent=2).encode("utf-8")

#             st.subheader("📥 Download Results")
#             st.download_button("⬇️ Download Excel", excel_data,
#                                file_name=f"{company_name}_normalized.xlsx",
#                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#             st.download_button("⬇️ Download CSV", csv_data,
#                                file_name=f"{company_name}_normalized.csv", mime="text/csv")
#             st.download_button("⬇️ Download JSON", json_data,
#                                file_name=f"{company_name}_normalized.json", mime="application/json")

#     except Exception as e:
#         st.error(f"❌ Error: {e}")


import os
import re
import json
import base64
import pandas as pd
from io import BytesIO
import streamlit as st
from dotenv import load_dotenv
from openai import OpenAI

# Load API Key
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    st.error("⚠️ OPENAI_API_KEY not set in .env or environment variables.")
    st.stop()

client = OpenAI(api_key=OPENAI_API_KEY)

# ---- Prompt Template ----
PROMPT_TEMPLATE = """
You are given insurance policy text or tabular data.
Extract and normalize into JSON with these fields:
- Segment
- Location
- Policy Type
- Doable Districts
- Pay in
- Transaction
- Age
- Remarks
- Insurer

Rules:
- If Location is present, put it in "Location".
- If RTO state is mentioned, add it into "Remarks" as "RTO State: <state>".
- Doable Districts means the RTOs or places mentioned.
- Policy Type may appear as "Types", "Type Comp/TP", "AOTP", "All" — keep it EXACTLY as written in the data.
- If Policy Type is missing, set it as "COMP/TP".
- Preserve duplicates: if multiple fields with same name exist, capture all.
- Headers may sometimes be in rows or sideways — detect intelligently.
- Keep ALL extra info (vehicle make, age, validity, etc.) in Remarks.
- Set Insurer to the provided company name: {company_name}.

Return only a JSON array.
Text:
{text}
"""

def clean_json_response(response_text: str) -> str:
    """Remove markdown formatting and keep pure JSON"""
    return re.sub(r"```json|```", "", response_text).strip()

def parse_transposed_excel(df: pd.DataFrame, company_name: str) -> pd.DataFrame:
    """Handle reports where first column contains field names, and each other column is a record."""
    df = df.dropna(how="all").reset_index(drop=True)
    fields = df.iloc[:, 0].astype(str).str.strip().tolist()

    records = []
    for col in df.columns[1:]:
        values = df[col].tolist()
        record = dict(zip(fields, values))

        rec = {
            "Segment": str(record.get("Segment", "")).strip() or "N/A",
            "Location": str(record.get("RTO state", "")).strip() or "N/A",
            "Policy Type": str(record.get("Type Comp/TP", record.get("Type", ""))).strip() or "COMP/TP",
            "Doable Districts": str(record.get("Doable District Name", "")).strip() or "N/A",
            "Pay in": str(record.get("Revised PO", "")).strip() or "N/A",
            "Transaction": str(record.get("Transaction New/Old", "")).strip() or "N/A",
            "Age": str(record.get("Age", "")).strip() or "N/A",
            "Remarks": str(record.get("Any Special remarks", "")).strip() or "N/A",
            "Insurer": company_name
        }
        records.append(rec)

    return pd.DataFrame(records)

def extract_excel(file_bytes: bytes, company_name: str) -> pd.DataFrame:
    """Extract and normalize Excel (supports transposed layout and duplicates)."""
    xls = pd.ExcelFile(BytesIO(file_bytes))
    all_records = []

    for sheet in xls.sheet_names:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=None, mangle_dupe_cols=False)

        first_col = df.iloc[:, 0].astype(str).str.lower().tolist()
        if any("segment" in v for v in first_col):
            parsed = parse_transposed_excel(df, company_name)
            parsed["Source_Sheet"] = sheet
            all_records.append(parsed)
        else:
            df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, mangle_dupe_cols=False)
            df["Insurer"] = company_name
            df["Source_Sheet"] = sheet
            all_records.append(df)

    return pd.concat(all_records, ignore_index=True)

def extract_text(file_bytes: bytes, filename: str, company_name: str):
    """Extracts text from different file formats"""
    ext = filename.split(".")[-1].lower()

    if ext in ["xlsx", "xls"]:
        return extract_excel(file_bytes, company_name)

    if ext == "csv":
        df = pd.read_csv(BytesIO(file_bytes), mangle_dupe_cols=False)
        df["Insurer"] = company_name
        return df

    if ext == "txt":
        return file_bytes.decode("utf-8", errors="ignore")

    if ext == "pdf":
        import pdfplumber
        text = ""
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
        return text

    if ext in ["jpg", "jpeg", "png"]:
        b64 = base64.b64encode(file_bytes).decode("utf-8")
        resp = client.chat.completions.create(
            model="gpt-4o",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": "Extract insurance policy text accurately"},
                    {"type": "image_url", "image_url": {"url": f"data:image/{ext};base64,{b64}"}}
                ]
            }],
            temperature=0.1
        )
        return resp.choices[0].message.content

    return ""

def process_file(file_bytes: bytes, filename: str, company_name: str):
    """Process the uploaded file and return a dataframe"""
    result = extract_text(file_bytes, filename, company_name)

    if isinstance(result, pd.DataFrame):
        return result

    # For non-dataframe results (txt, pdf, images), use AI parsing
    prompt = PROMPT_TEMPLATE.format(text=result[:12000], company_name=company_name)

    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1
    )
    result = clean_json_response(resp.choices[0].message.content)

    try:
        records = json.loads(result)
    except:
        records = [{
            "Segment": "N/A",
            "Location": "N/A",
            "Policy Type": "COMP/TP",
            "Doable Districts": "N/A",
            "Pay in": "N/A",
            "Transaction": "N/A",
            "Age": "N/A",
            "Remarks": result,
            "Insurer": company_name
        }]

    for rec in records:
        pt = rec.get("Policy Type", "").strip()
        if not pt or pt in ["", "N/A"]:
            rec["Policy Type"] = "COMP/TP"
        rec["Insurer"] = company_name

    return pd.DataFrame(records)

def make_excel(df: pd.DataFrame, company_name: str) -> bytes:
    """Create Excel file with company name as header"""
    from openpyxl.styles import Alignment, Font
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Normalized Data", startrow=2, index=False)
        ws = writer.sheets["Normalized Data"]

        headers = list(df.columns)
        for col_num, value in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=value)
            cell.font = Font(bold=True)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=company_name).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1, value="Normalized Insurance Policy Data").font = Font(bold=True, size=12)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    output.seek(0)
    return output.read()

# ---- Streamlit UI ----
st.set_page_config(page_title="Insurance Report Normalizer", page_icon="📊", layout="wide")
st.title("📊 Insurance Report Normalizer")
st.write("Upload your insurance report (PDF, Excel, CSV, TXT, Image) to convert into a structured Excel sheet.")

with st.sidebar:
    st.header("Settings")
    company_name = st.text_input("🏢 Company / Insurer Name", value="Unknown Company")
    uploaded_file = st.file_uploader(
        "📂 Upload Report File",
        type=["pdf", "xlsx", "xls", "csv", "txt", "jpg", "jpeg", "png"]
    )

if uploaded_file and company_name:
    st.info(f"Processing file: {uploaded_file.name} ...")
    try:
        df = process_file(uploaded_file.read(), uploaded_file.name, company_name)
        st.success("✅ File processed successfully!")
        st.subheader("Preview of Normalized Data")
        st.dataframe(df, use_container_width=True)

        excel_data = make_excel(df, company_name)
        csv_data = df.to_csv(index=False).encode("utf-8")
        json_data = df.to_json(orient="records", indent=2).encode("utf-8")

        st.subheader("📥 Download Results")
        st.download_button("⬇️ Download Excel", excel_data,
                           file_name=f"{company_name}_normalized.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.download_button("⬇️ Download CSV", csv_data,
                           file_name=f"{company_name}_normalized.csv", mime="text/csv")
        st.download_button("⬇️ Download JSON", json_data,
                           file_name=f"{company_name}_normalized.json", mime="application/json")

    except Exception as e:
        st.error(f"❌ Error: {e}")
